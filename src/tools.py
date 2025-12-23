import asyncio
import os
from datetime import datetime
from typing import List, Dict, Optional, Literal

from langchain_mcp_adapters.client import MultiServerMCPClient
from langchain.agents import create_agent

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    openpyxl = None

try:
    import json
    from pathlib import Path
except ImportError:
    json = None
    Path = None

def get_weather(city:str) ->str:
    """get the weather for a city"""
    return f"{city}.今天是晴天，温度为25度."

def get_zhipu_search_mcp_tools():
    client = MultiServerMCPClient(
        {
            "search": {
                "transport": "sse",  # HTTP-based remote server
                # Ensure you start your weather server on port 8000
                "url": "https://open.bigmodel.cn/api/mcp/web_search/sse?Authorization=dfccd65ce1c9466690f465304cb6ae09.KzH9GpjmAVgN4Jhr",
            }
        }
    )
    tools = asyncio.run(client.get_tools())
    return tools
def get_tavily_search_mcp_tools():
    client = MultiServerMCPClient(
        {
            "search": {
                "transport": "streamable_http",  # HTTP-based remote server
                # Ensure you start your weather server on port 8000
                "url": "https://mcp.tavily.com/mcp/?tavilyApiKey=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",
            }
        }
    )
    tools = asyncio.run(client.get_tools())
    return tools
def get_chrome_mcp_tools():
    client = MultiServerMCPClient(
        {
            "chrome_mcp": {
                "transport": "streamable_http",  # HTTP-based remote server
                # Ensure you start your weather server on port 8000
                "url": "http://127.0.0.1:12306/mcp",
            }
        }
    )
    tools = asyncio.run(client.get_tools())
    return tools
def get_mcp_server_chart_tools():
    client = MultiServerMCPClient(
        {
            "mcp_chart_server": {
                "command":"cmd",
                "args": ["/c", "npx", "-y", "@antv/mcp-server-chart"],
                "transport": "stdio",  # HTTP-based remote server
            }
        }
    )
    tools = asyncio.run(client.get_tools())
    return tools
# tools = get_zhipu_search_mcp_tools()
# print(tools)
# os.environ["TAVILY_API_KEY"] = "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"
# from langchain_tavily import TavilySearch
# toolSearch = TavilySearch(max_results=2)


def save_test_cases_to_excel(
    test_cases: List[Dict],
    file_path: str = "test_cases.xlsx",
    sheet_name: str = "测试用例",
    mode: Literal["overwrite", "append"] = "overwrite",
) -> str:
    """
    将测试用例保存到 Excel 文件中

    Args:
        test_cases: 测试用例列表，每个测试用例是一个字典
                   示例: [
                       {
                           "case_id": "TC001",
                           "title": "登录功能测试",
                           "precondition": "用户未登录",
                           "steps": "1. 打开登录页\n2. 输入用户名密码\n3. 点击登录",
                           "expected": "登录成功",
                           "priority": "高",
                           "status": "通过"
                       }
                   ]
        file_path: Excel 文件保存路径，默认为 "test_cases.xlsx"
        sheet_name: 工作表名称，默认为 "测试用例"

    Returns:
        str: 保存结果信息
    """
    # 检查是否安装了 openpyxl 库
    if openpyxl is None:
        return "错误: 未安装 openpyxl 库，请运行: pip install openpyxl"

    # 检查测试用例列表是否为空
    if not test_cases:
        return "错误: 测试用例列表为空"

    try:
        file_exists = os.path.exists(file_path)
        if file_exists:
            workbook = openpyxl.load_workbook(file_path)
        else:
            workbook = openpyxl.Workbook()

        # 获取/创建工作表
        if mode == "overwrite" and sheet_name in workbook.sheetnames:
            old_sheet = workbook[sheet_name]
            workbook.remove(old_sheet)

        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.create_sheet(title=sheet_name)

        # openpyxl 默认新建会带一个 Sheet，这里尽量保持整洁
        if "Sheet" in workbook.sheetnames and len(workbook.sheetnames) > 1:
            default_sheet = workbook["Sheet"]
            if default_sheet.max_row == 1 and default_sheet.max_column == 1 and default_sheet["A1"].value is None:
                workbook.remove(default_sheet)

        # mode=append 时，尽量复用已有表头；否则从 test_cases 生成表头
        existing_headers: list[str] = []
        if mode == "append" and sheet.max_row >= 1:
            first_row = [c.value for c in sheet[1]]
            if any(v is not None for v in first_row):
                existing_headers = [str(v) for v in first_row if v is not None]

        headers = existing_headers or list(test_cases[0].keys())

        # 补齐新增列（append 且 test_cases 出现新字段）
        if existing_headers:
            extra = [k for k in test_cases[0].keys() if k not in existing_headers]
            if extra:
                headers = existing_headers + extra

        # 设置表头样式
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        def write_header_row() -> None:
            for col_idx, header in enumerate(headers, start=1):
                cell = sheet.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border

        def header_is_empty() -> bool:
            if sheet.max_row < 1:
                return True
            row = [c.value for c in sheet[1]]
            return not any(v is not None and str(v).strip() for v in row)

        if mode == "overwrite" or header_is_empty():
            sheet.delete_rows(1, sheet.max_row)
            write_header_row()
        else:
            # 若 append 且出现新列，更新表头行
            if existing_headers and headers != existing_headers:
                write_header_row()

        start_row = sheet.max_row + 1 if (mode == "append" and sheet.max_row >= 2) else 2

        # 写入测试用例数据
        for row_idx, test_case in enumerate(test_cases, start=start_row):
            for col_idx, header in enumerate(headers, start=1):
                # 获取对应字段的值
                value = test_case.get(header, "")
                cell = sheet.cell(row=row_idx, column=col_idx, value=value)

                # 设置单元格样式
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = border

                # 根据状态设置不同颜色（如果有 status 字段）
                if header.lower() in ["status", "状态"]:
                    if value in ["通过", "Pass", "PASS"]:
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    elif value in ["失败", "Fail", "FAIL"]:
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    elif value in ["阻塞", "Blocked", "BLOCKED"]:
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

        # 自动调整列宽（避免 append 时全表扫描，按本次写入的行段扫描）
        scan_start = 1 if mode == "overwrite" else max(1, start_row - 1)
        scan_end = sheet.max_row
        for col_idx, header in enumerate(headers, start=1):
            max_length = len(str(header))
            for row_idx in range(scan_start, scan_end + 1):
                cell_value = str(sheet.cell(row=row_idx, column=col_idx).value or "")
                max_line_length = max((len(line) for line in cell_value.split("\n")), default=0)
                max_length = max(max_length, max_line_length)

            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = adjusted_width

        # 冻结首行（表头）
        sheet.freeze_panes = "A2"

        # 保存文件
        workbook.save(file_path)

        # 返回成功信息
        return f"成功: 已将 {len(test_cases)} 条测试用例保存到 {os.path.abspath(file_path)}"

    except Exception as e:
        return f"错误: 保存失败 - {str(e)}"


def save_and_generate_report(
    test_cases: List[Dict],
    chart_html: str = "",
    file_path: str = "test_report.html",
    report_title: str = "UI自动化测试报告"
) -> str:
    """
    生成包含图表的HTML测试报告并保存到指定路径

    Args:
        test_cases: 测试用例列表
        chart_html: 图表的HTML代码（从mcp-server-chart生成）
        file_path: 报告保存路径
        report_title: 报告标题

    Returns:
        str: 保存结果信息
    """
    if not test_cases:
        return "错误: 测试用例列表为空"

    try:
        # 统计数据
        total = len(test_cases)
        passed = sum(1 for tc in test_cases if tc.get("状态") in ["通过", "Pass", "PASS"] or tc.get("status") in ["通过", "Pass", "PASS"])
        failed = sum(1 for tc in test_cases if tc.get("状态") in ["失败", "Fail", "FAIL"] or tc.get("status") in ["失败", "Fail", "FAIL"])
        blocked = sum(1 for tc in test_cases if tc.get("状态") in ["阻塞", "Blocked", "BLOCKED"] or tc.get("status") in ["阻塞", "Blocked", "BLOCKED"])
        pass_rate = (passed / total * 100) if total > 0 else 0

        # 生成测试用例表格HTML
        table_rows = ""
        for tc in test_cases:
            status = tc.get("状态") or tc.get("status", "")
            status_class = "pass" if status in ["通过", "Pass", "PASS"] else "fail" if status in ["失败", "Fail", "FAIL"] else "blocked"
            table_rows += f"""
            <tr>
                <td>{tc.get("用例ID") or tc.get("case_id", "")}</td>
                <td>{tc.get("用例标题") or tc.get("title", "")}</td>
                <td>{tc.get("测试步骤") or tc.get("steps", "")}</td>
                <td>{tc.get("预期结果") or tc.get("expected", "")}</td>
                <td>{tc.get("实际结果") or tc.get("actual", "")}</td>
                <td class="{status_class}">{status}</td>
            </tr>
            """

        # HTML模板
        html_content = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{report_title}</title>
    <style>
        body {{ font-family: Arial, sans-serif; margin: 20px; background: #f5f5f5; }}
        .container {{ max-width: 1200px; margin: 0 auto; background: white; padding: 30px; border-radius: 8px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }}
        h1 {{ color: #333; border-bottom: 3px solid #4472C4; padding-bottom: 10px; }}
        .summary {{ display: flex; gap: 20px; margin: 20px 0; }}
        .stat-card {{ flex: 1; padding: 20px; border-radius: 8px; text-align: center; }}
        .stat-card h3 {{ margin: 0; font-size: 32px; }}
        .stat-card p {{ margin: 5px 0 0 0; color: #666; }}
        .total {{ background: #E7F3FF; color: #0066CC; }}
        .pass {{ background: #E8F5E9; color: #2E7D32; }}
        .fail {{ background: #FFEBEE; color: #C62828; }}
        .blocked {{ background: #FFF9E6; color: #F57C00; }}
        .chart-section {{ margin: 30px 0; padding: 20px; background: #fafafa; border-radius: 8px; }}
        table {{ width: 100%; border-collapse: collapse; margin: 20px 0; }}
        th, td {{ padding: 12px; text-align: left; border: 1px solid #ddd; }}
        th {{ background: #4472C4; color: white; font-weight: bold; }}
        tr:nth-child(even) {{ background: #f9f9f9; }}
        .timestamp {{ color: #666; font-size: 14px; margin-top: 20px; text-align: right; }}
    </style>
</head>
<body>
    <div class="container">
        <h1>{report_title}</h1>

        <div class="summary">
            <div class="stat-card total">
                <h3>{total}</h3>
                <p>总用例数</p>
            </div>
            <div class="stat-card pass">
                <h3>{passed}</h3>
                <p>通过</p>
            </div>
            <div class="stat-card fail">
                <h3>{failed}</h3>
                <p>失败</p>
            </div>
            <div class="stat-card blocked">
                <h3>{blocked}</h3>
                <p>阻塞</p>
            </div>
        </div>

        <div class="stat-card" style="background: #F0F4FF; margin: 20px 0;">
            <h3>{pass_rate:.1f}%</h3>
            <p>通过率</p>
        </div>

        {f'<div class="chart-section"><h2>测试统计图表</h2>{chart_html}</div>' if chart_html else ''}

        <h2>测试用例详情</h2>
        <table>
            <thead>
                <tr>
                    <th>用例ID</th>
                    <th>用例标题</th>
                    <th>测试步骤</th>
                    <th>预期结果</th>
                    <th>实际结果</th>
                    <th>状态</th>
                </tr>
            </thead>
            <tbody>
                {table_rows}
            </tbody>
        </table>

        <div class="timestamp">生成时间: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}</div>
    </div>
</body>
</html>"""

        # 保存文件
        Path(file_path).parent.mkdir(parents=True, exist_ok=True)
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(html_content)

        return f"成功: 测试报告已保存到 {os.path.abspath(file_path)}"

    except Exception as e:
        return f"错误: 生成报告失败 - {str(e)}"


# =====================================================================
# PDF 附件解析工具（供 src/file_rag/main.py 的 before_model 中间件调用）
# =====================================================================
#
# 背景说明（新手友好版）：
# - 前端上传 PDF 时，messages 里会带一个 {"type": "file", "source_type": "base64", ...} 的块
# - 如果把 base64 直接塞给模型，通常会出现：
#   1) 请求体太大（卡顿/超时/400）
#   2) 模型不认识 "file" block（400 unknown variant file）
#
# 所以这里做“预处理”：
# 1) base64 -> 真正的 PDF bytes
# 2) 用 langchain-pymupdf4llm 抽取 PDF 文本
# 3) 在需要时，调用多模态模型把 PDF 图片转成文本（可检索、可问答）
# 4) 最终把解析出来的文本回写到 HumanMessage.content（让模型看到“文本”而不是 base64）
#
# 注意：
# - 多模态解析图片开销很大，这里默认“按需启用” + “预算限制”，避免前端看起来卡住
#

import base64
import gc
import hashlib
import json
import re
import time
from tempfile import TemporaryDirectory
from typing import Any
from uuid import uuid4

from langchain_core.documents import Document
from langchain_core.document_loaders import BaseBlobParser, Blob
from langchain_core.messages import HumanMessage, SystemMessage
from langchain_pymupdf4llm import PyMuPDF4LLMLoader, PyMuPDF4LLMParser

from langchain.tools import tool

from file_rag.core.llms import get_default_model, get_doubao_seed_model


def _env_int(name: str, default: int) -> int:
    """读取 int 环境变量，失败则返回 default。"""
    try:
        return int(os.getenv(name, str(default)))
    except Exception:
        return default


def _env_float(name: str, default: float) -> float:
    """读取 float 环境变量，失败则返回 default。"""
    try:
        return float(os.getenv(name, str(default)))
    except Exception:
        return default


# ==========================
# 可调参数（性能/体验相关）
# ==========================
# 你可以通过环境变量调整（不需要改代码）：
# - PDF_FORCE_EXTRACT_IMAGES=1  强制开启图片多模态解析（无论用户问题是否提到图片）
# - PDF_IMAGE_MAX_PAGES=5       最多解析前 N 页的图片（默认 5）
# - PDF_IMAGE_MAX_IMAGES=10     最多解析前 N 张图片（默认 10）
# - PDF_IMAGE_MAX_SECONDS=20    图片解析总耗时上限（秒，默认 20s）
# - PDF_IMAGE_MAX_BYTES=2500000 单张图片最大 bytes（默认 2.5MB）
# - PDF_TEXT_MAX_CHARS=30000    注入到对话的文本最大字符数（默认 30000）
# - PDF_TEXT_MAX_PAGES=20       文本最多解析前 N 页（默认 20）
#
# 建议：
# - 如果你前端经常“转圈很久”，先把 PDF_TEXT_MAX_CHARS/页数上限调小，保证交互先跑通
#
_PDF_FORCE_EXTRACT_IMAGES: bool = os.getenv("PDF_FORCE_EXTRACT_IMAGES", "0").lower() in {
    "1",
    "true",
    "yes",
    "y",
}
_PDF_IMAGE_MAX_PAGES: int = _env_int("PDF_IMAGE_MAX_PAGES", 5)
_PDF_IMAGE_MAX_IMAGES: int = _env_int("PDF_IMAGE_MAX_IMAGES", 10)
_PDF_IMAGE_MAX_SECONDS: float = _env_float("PDF_IMAGE_MAX_SECONDS", 20.0)
_PDF_IMAGE_MAX_BYTES: int = _env_int("PDF_IMAGE_MAX_BYTES", 2_500_000)
_PDF_TEXT_MAX_CHARS: int = _env_int("PDF_TEXT_MAX_CHARS", 30000)
_PDF_TEXT_MAX_PAGES: int = _env_int("PDF_TEXT_MAX_PAGES", 20)

# 图片解析模式：
# - always：只要检测到 PDF 含图片/流程图等视觉内容，就尝试解析（仍受 PDF_IMAGE_MAX_* 预算限制）
# - auto：用户问题提到“图片/图表/流程图”等或检测到 PDF 含视觉内容时才解析
# - never：从不解析图片
_PDF_IMAGES_MODE: str = (os.getenv("PDF_IMAGES_MODE", "always") or "always").strip().lower()

# 是否把“原始 PDF 文件”持久化保存到本地（默认开启，避免信息丢失）
_PDF_PERSIST_UPLOADS: bool = os.getenv("PDF_PERSIST_UPLOADS", "1").lower() in {
    "1",
    "true",
    "yes",
    "y",
}

# 是否把“解析结果”按页分片保存（jsonl）（默认开启，便于后续做检索/RAG）
_PDF_STORE_CHUNKS: bool = os.getenv("PDF_STORE_CHUNKS", "1").lower() in {
    "1",
    "true",
    "yes",
    "y",
}

# 是否覆盖已有的分片文件（同一份 PDF 重复上传时可能命中同一个 doc_id）
_PDF_OVERWRITE_EXTRACTED: bool = os.getenv("PDF_OVERWRITE_EXTRACTED", "0").lower() in {
    "1",
    "true",
    "yes",
    "y",
}

_PDF_UPLOAD_DIR = os.getenv("PDF_UPLOAD_DIR", "storage/pdf_uploads")
_PDF_EXTRACT_DIR = os.getenv("PDF_EXTRACT_DIR", "storage/pdf_extracted")

# 是否复用已落盘的抽取结果（chunks/meta）（默认开启，避免重复解析/重复写入导致“从头再来”）
_PDF_REUSE_EXTRACTED: bool = os.getenv("PDF_REUSE_EXTRACTED", "1").lower() in {
    "1",
    "true",
    "yes",
    "y",
}

# 读取分片时是否去重（防止历史重复写入导致“看起来又从头解析一遍”）
_PDF_DEDUP_CHUNKS: bool = os.getenv("PDF_DEDUP_CHUNKS", "1").lower() in {
    "1",
    "true",
    "yes",
    "y",
}

# “注入到对话上下文”的预算（为了不卡住、避免上下文爆炸）
_PDF_CONTEXT_MAX_PAGES: int = _env_int("PDF_CONTEXT_MAX_PAGES", _PDF_TEXT_MAX_PAGES)
_PDF_CONTEXT_MAX_CHARS: int = _env_int("PDF_CONTEXT_MAX_CHARS", _PDF_TEXT_MAX_CHARS)

# “落盘分片抽取”的预算（用于存储原始解析结果，默认不限制，避免信息丢失）
_PDF_EXTRACT_MAX_PAGES: int = _env_int("PDF_EXTRACT_MAX_PAGES", 0)
_PDF_EXTRACT_MAX_CHARS: int = _env_int("PDF_EXTRACT_MAX_CHARS", 0)

# 落盘 jsonl 的单条最大字符数（把“按页”再切小一点，便于断点续跑）
_PDF_CHUNK_MAX_CHARS: int = _env_int("PDF_CHUNK_MAX_CHARS", 4000)


def _user_requested_images(user_text: str) -> bool:
    """判断用户是否明确要求“解析图片/流程图/图表”等视觉内容。"""
    return any(
        kw in (user_text or "")
        for kw in (
            "图片",
            "图中",
            "图里",
            "截图",
            "流程图",
            "示意图",
            "图表",
            "表格",
        )
    )


def _compute_pdf_id(pdf_bytes: bytes) -> str:
    """为一份 PDF 生成稳定的 id（用于去重、持久化目录命名）。"""
    return hashlib.sha256(pdf_bytes).hexdigest()


def _persist_pdf_upload(pdf_bytes: bytes, filename: str) -> tuple[str | None, Path | None]:
    """把原始 PDF bytes 落盘保存，避免“预算限制”导致信息不可恢复。

    返回：
    - (doc_id, pdf_path)
      - doc_id: sha256(pdf_bytes)
      - pdf_path: 保存后的真实路径

    说明：
    - 如果关闭了 PDF_PERSIST_UPLOADS，则返回 (None, None)
    """
    if not _PDF_PERSIST_UPLOADS:
        return None, None

    doc_id = _compute_pdf_id(pdf_bytes)
    base_dir = Path(_PDF_UPLOAD_DIR) / doc_id
    base_dir.mkdir(parents=True, exist_ok=True)

    pdf_path = base_dir / _safe_pdf_filename(filename)
    if not pdf_path.exists():
        pdf_path.write_bytes(pdf_bytes)

    meta_path = base_dir / "meta.json"
    if _PDF_OVERWRITE_EXTRACTED or not meta_path.exists():
        meta = {
            "doc_id": doc_id,
            "filename": _safe_pdf_filename(filename),
            "bytes": len(pdf_bytes),
            "pdf_path": pdf_path.as_posix(),
            "created_at": datetime.now().isoformat(timespec="seconds"),
        }
        meta_path.write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8")

    return doc_id, pdf_path


def _open_chunks_writer(doc_id: str, *, filename: str, pdf_path: Path) -> tuple[Path, Any] | tuple[None, None]:
    """打开 jsonl 分片写入器：每行一个 chunk（按页/按阶段）。"""
    if not _PDF_STORE_CHUNKS:
        return None, None

    base_dir = Path(_PDF_EXTRACT_DIR) / doc_id
    base_dir.mkdir(parents=True, exist_ok=True)

    chunks_path = base_dir / "chunks.jsonl"
    if _PDF_OVERWRITE_EXTRACTED and chunks_path.exists():
        chunks_path.write_text("", encoding="utf-8")
    elif chunks_path.exists() and _PDF_REUSE_EXTRACTED:
        # 关键：如果已经存在抽取结果，默认不再“追加写入”，避免：
        # - chunks.jsonl 被重复写入同样内容（导致分析阶段看起来“反复从头解析”）
        # - 文件越来越大，后续分析耗时越来越长
        return chunks_path, None

    meta_path = base_dir / "meta.json"
    if _PDF_OVERWRITE_EXTRACTED or not meta_path.exists():
        meta = {
            "doc_id": doc_id,
            "filename": _safe_pdf_filename(filename),
            "pdf_path": pdf_path.as_posix(),
            "budgets": {
                "PDF_CONTEXT_MAX_PAGES": _PDF_CONTEXT_MAX_PAGES,
                "PDF_CONTEXT_MAX_CHARS": _PDF_CONTEXT_MAX_CHARS,
                "PDF_EXTRACT_MAX_PAGES": _PDF_EXTRACT_MAX_PAGES,
                "PDF_EXTRACT_MAX_CHARS": _PDF_EXTRACT_MAX_CHARS,
                "PDF_IMAGE_MAX_PAGES": _PDF_IMAGE_MAX_PAGES,
                "PDF_IMAGE_MAX_IMAGES": _PDF_IMAGE_MAX_IMAGES,
                "PDF_IMAGE_MAX_SECONDS": _PDF_IMAGE_MAX_SECONDS,
                "PDF_IMAGE_MAX_BYTES": _PDF_IMAGE_MAX_BYTES,
                "PDF_IMAGES_MODE": _PDF_IMAGES_MODE,
            },
            "created_at": datetime.now().isoformat(timespec="seconds"),
        }
        meta_path.write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8")

    # 新文件：写入分片（默认只写一次；后续命中同 doc_id 则复用，不再追加）
    fp = open(chunks_path, "a", encoding="utf-8")
    return chunks_path, fp


def _extracted_ready(doc_id: str) -> bool:
    """判断 doc_id 是否已经有可复用的落盘抽取结果。"""
    base_dir = Path(_PDF_EXTRACT_DIR) / doc_id
    return (base_dir / "chunks.jsonl").exists() and (base_dir / "meta.json").exists()


def _split_text(text: str, max_chars: int) -> list[str]:
    """把长文本切成多个小块（用于 jsonl 分片与断点续跑）。"""
    if not isinstance(text, str) or not text:
        return [""]
    if max_chars <= 0 or len(text) <= max_chars:
        return [text]
    parts: list[str] = []
    start = 0
    while start < len(text):
        end = min(len(text), start + max_chars)
        parts.append(text[start:end])
        start = end
    return parts


def _build_context_excerpt_from_chunks(doc_id: str) -> str:
    """从已落盘的 chunks.jsonl 构造“注入上下文”的节选，避免重复解析 PDF。"""
    chunks_path = Path(_PDF_EXTRACT_DIR) / doc_id / "chunks.jsonl"
    if not chunks_path.exists():
        return ""

    pages_seen: set[int] = set()
    pages_count = 0
    chars = 0
    out_parts: list[str] = []

    with open(chunks_path, "r", encoding="utf-8") as fp:
        for line in fp:
            line = line.strip()
            if not line:
                continue
            try:
                obj = json.loads(line)
            except Exception:
                continue
            if not isinstance(obj, dict):
                continue

            content = obj.get("content", "")
            if not isinstance(content, str) or not content.strip():
                continue

            page = obj.get("page")
            if isinstance(page, int):
                if page not in pages_seen:
                    pages_seen.add(page)
                    pages_count += 1
                if _PDF_CONTEXT_MAX_PAGES > 0 and pages_count > _PDF_CONTEXT_MAX_PAGES:
                    break

            if _PDF_CONTEXT_MAX_CHARS > 0 and chars >= _PDF_CONTEXT_MAX_CHARS:
                break

            remaining = _PDF_CONTEXT_MAX_CHARS - chars if _PDF_CONTEXT_MAX_CHARS > 0 else None
            if remaining is not None and remaining <= 0:
                break

            text_to_add = content if remaining is None else content[:remaining]
            out_parts.append(text_to_add)
            chars += len(text_to_add)

    excerpt = "\n\n".join([p for p in out_parts if p.strip()]).strip()
    if excerpt:
        excerpt += (
            f"\n\n[提示：以上为落盘分片的节选（DOC_ID: {doc_id}）。"
            f"需要“完整不漏”的分析请使用 pdf_analyze_doc(DOC_ID, question)。]"
        )
    return excerpt


def _truncate_text(text: str, max_chars: int) -> tuple[str, bool]:
    """把超长文本截断到 max_chars，返回 (截断后的文本, 是否截断过)。"""
    if max_chars <= 0:
        return text, False
    if len(text) <= max_chars:
        return text, False
    return (
        text[:max_chars].rstrip()
        + f"\n\n[内容过长已截断：仅保留前 {max_chars} / {len(text)} 字符，可通过 PDF_TEXT_MAX_CHARS 调整上限]",
        True,
    )


_IMAGE_TO_TEXT_PROMPT: str = (
    "你是一个用于需求/PRD 多模态解析的助手，任务是把图片内容转成【可检索、可追溯、可用于生成测试用例】的文本。\n"
    "\n"
    "输出要求（只输出 Markdown，不要包含解释性文字，也不要在开头输出```）：\n"
    "1) 先输出“关键信息摘要”（<= 12 条，信息密度高）：包含页面/模块、角色/权限、输入输出、关键字段、规则/阈值、异常分支。\n"
    "2) 再输出“完整文字转录”：按从上到下、从左到右顺序尽量不漏。\n"
    "3) 若图片是流程图/状态机/时序图/泳道图：必须输出“流程/分支清单”，用 `A -> B（条件/事件：...）` 形式列出所有可见分支/回路/终止条件。\n"
    "4) 若图片是界面原型/截图：必须输出“交互要点”，列出按钮/入口、校验提示、错误码/异常提示文案、可选项/默认值。\n"
)


def _try_get_image_size(raw: bytes, mimetype: str) -> tuple[int | None, int | None]:
    """尽量从图片 bytes 中解析出宽高（不依赖 Pillow）。"""
    try:
        mt = (mimetype or "").lower()

        # ---- PNG ----
        if mt == "image/png":
            if len(raw) < 24:
                return None, None
            if raw[:8] != b"\x89PNG\r\n\x1a\n":
                return None, None
            if raw[12:16] != b"IHDR":
                return None, None
            w = int.from_bytes(raw[16:20], "big")
            h = int.from_bytes(raw[20:24], "big")
            return (w, h) if w > 0 and h > 0 else (None, None)

        # ---- JPEG ----
        if mt in {"image/jpeg", "image/jpg"}:
            if len(raw) < 4 or raw[0:2] != b"\xFF\xD8":
                return None, None
            i = 2
            n = len(raw)
            while i + 9 < n:
                if raw[i] != 0xFF:
                    i += 1
                    continue
                marker = raw[i + 1]
                i += 2
                if marker in {0xC0, 0xC2}:
                    if i + 7 >= n:
                        break
                    seg_len = int.from_bytes(raw[i : i + 2], "big")
                    if seg_len < 7:
                        break
                    h = int.from_bytes(raw[i + 3 : i + 5], "big")
                    w = int.from_bytes(raw[i + 5 : i + 7], "big")
                    return (w, h) if w > 0 and h > 0 else (None, None)
                if i + 1 >= n:
                    break
                seg_len = int.from_bytes(raw[i : i + 2], "big")
                i += seg_len
            return None, None

        # ---- WEBP ----
        if mt == "image/webp":
            if len(raw) < 30:
                return None, None
            if not (raw.startswith(b"RIFF") and raw[8:12] == b"WEBP"):
                return None, None
            i = 12
            n = len(raw)
            while i + 8 <= n:
                tag = raw[i : i + 4]
                size = int.from_bytes(raw[i + 4 : i + 8], "little")
                i += 8
                if i + size > n:
                    break
                if tag == b"VP8X" and size >= 10 and i + 10 <= n:
                    w = int.from_bytes(raw[i + 4 : i + 7], "little") + 1
                    h = int.from_bytes(raw[i + 7 : i + 10], "little") + 1
                    return (w, h) if w > 0 and h > 0 else (None, None)
                i += size + (size % 2)
            return None, None

        return None, None
    except Exception:
        return None, None


class _BudgetedImageBlobParser(BaseBlobParser):
    """给图片多模态解析加“预算限制”，避免图片太多导致长时间阻塞。"""

    def __init__(self, inner: BaseBlobParser, *, max_images: int) -> None:
        super().__init__()
        self._inner = inner
        self._max_images = max_images
        self.parsed_images = 0
        self.skipped_images = 0

    def lazy_parse(self, blob: Blob):
        if self._max_images >= 0 and self.parsed_images >= self._max_images:
            self.skipped_images += 1
            yield Document(
                page_content=f"[图片过多，已跳过多模态解析：max_images={self._max_images}]",
                metadata={**blob.metadata, **{"source": blob.source}},
            )
            return

        self.parsed_images += 1
        yield from self._inner.lazy_parse(blob)


class _MultimodalImageBlobParser(BaseBlobParser):
    """把图片 Blob 交给多模态模型，输出图片描述/图片文字。"""

    def __init__(self, *, model, prompt: str = _IMAGE_TO_TEXT_PROMPT) -> None:
        super().__init__()
        self._model = model
        self._prompt = prompt
        # doubao 触发 429 时，后续图片不再继续打请求（避免刷屏 + 避免前端看起来“卡住”）
        self._rate_limited = False
        self._rate_limit_notice_emitted = False

    def lazy_parse(self, blob: Blob):
        with blob.as_bytes_io() as buf:
            raw = buf.read()

        if _PDF_IMAGE_MAX_BYTES > 0 and len(raw) > _PDF_IMAGE_MAX_BYTES:
            mb = len(raw) / (1024 * 1024)
            content = (
                f"[图片过大（{mb:.2f}MB），已跳过多模态解析："
                f"PDF_IMAGE_MAX_BYTES={_PDF_IMAGE_MAX_BYTES}]"
            )
            yield Document(
                page_content=content,
                metadata={**blob.metadata, **{"source": blob.source}},
            )
            return

        mimetype = blob.mimetype or "image/png"
        if mimetype == "application/octet-stream" and blob.source:
            suffix = Path(str(blob.source)).suffix.lower()
            mimetype = {
                ".png": "image/png",
                ".jpg": "image/jpeg",
                ".jpeg": "image/jpeg",
                ".webp": "image/webp",
            }.get(suffix, "image/png")

        width, height = _try_get_image_size(raw, mimetype)
        if width is not None and height is not None and min(width, height) < 14:
            content = f"[图片尺寸过小（{width}x{height}），已跳过多模态解析]"
            yield Document(
                page_content=content,
                metadata={**blob.metadata, **{"source": blob.source}},
            )
            return

        img_base64 = base64.b64encode(raw).decode("utf-8")

        # 如果已经确认模型被限流/暂停，后续图片直接跳过（不再发请求）
        if self._rate_limited:
            if not self._rate_limit_notice_emitted:
                self._rate_limit_notice_emitted = True
                yield Document(
                    page_content=(
                        "[图片多模态解析已暂停：doubao 模型触发 429（Safe Experience Mode / SetLimitExceeded）。"
                        "请在火山引擎模型控制台调整/关闭 Safe Experience Mode 或提升额度后重试。]"
                    ),
                    metadata={**blob.metadata, **{"source": blob.source}},
                )
            else:
                # 返回空内容，避免把“重复提示”写进分片/上下文导致膨胀
                yield Document(
                    page_content="",
                    metadata={**blob.metadata, **{"source": blob.source}},
                )
            return

        try:
            msg = self._model.invoke(
                [
                    HumanMessage(
                        content=[
                            {"type": "text", "text": self._prompt},
                            {
                                "type": "image_url",
                                "image_url": {
                                    "url": f"data:{mimetype};base64,{img_base64}"
                                },
                            },
                        ]
                    )
                ]
            )
            content = msg.content if isinstance(msg.content, str) else str(msg.content)
        except Exception as exc:
            err = str(exc)
            # 识别“账号推理额度被暂停/限流”的典型报错，后续直接熔断，避免每张图片都打一遍 429。
            if any(
                k in err
                for k in (
                    "SetLimitExceeded",
                    "Safe Experience Mode",
                    "TooManyRequests",
                    "RateLimitError",
                    "429",
                )
            ):
                self._rate_limited = True
            if len(err) > 300:
                err = err[:300] + "…"
            content = f"[图片解析失败（已跳过该图片）：{err}]"

        yield Document(
            page_content=content,
            metadata={**blob.metadata, **{"source": blob.source}},
        )


_pdf_images_parser: BaseBlobParser | None = None
_pdf_images_parser_init_error: Exception | None = None


def _get_pdf_images_parser() -> BaseBlobParser | None:
    """懒加载 + 缓存图片解析器（失败后不反复重试）。"""
    global _pdf_images_parser, _pdf_images_parser_init_error

    if _pdf_images_parser is not None:
        return _pdf_images_parser
    if _pdf_images_parser_init_error is not None:
        return None

    try:
        _pdf_images_parser = _MultimodalImageBlobParser(model=get_doubao_seed_model())
        return _pdf_images_parser
    except Exception as exc:
        _pdf_images_parser_init_error = exc
        return None


def _decode_base64(data: str) -> bytes:
    """把 base64 字符串解码为 bytes（兼容 data URL）。"""
    raw = data.strip()
    if raw.startswith("data:") and "," in raw:
        raw = raw.split(",", 1)[1]
    raw = re.sub(r"\s+", "", raw)

    missing_padding = len(raw) % 4
    if missing_padding:
        raw += "=" * (4 - missing_padding)

    return base64.b64decode(raw)


def _safe_pdf_filename(filename: str | None) -> str:
    """生成安全文件名（避免路径穿越，缺省补 .pdf）。"""
    if not filename:
        return "upload.pdf"
    name = Path(filename).name
    return name if name.lower().endswith(".pdf") else f"{name}.pdf"


def _scan_pdf_visual_content(
    pdf_path: Path,
    *,
    max_pages: int,
    max_seconds: float,
) -> tuple[bool | None, list[int], int | None, bool]:
    """扫描 PDF 的视觉内容，用于决定是否启用豆包多模态解析。

    返回：
    - has_images：是否检测到嵌入位图（None 表示扫描失败/不可用）
    - render_pages：需要“页面渲染”解析的页号列表（0-based）
    - total_pages：总页数（可能为 None）
    - timed_out：是否因 max_seconds 提前停止扫描
    """
    try:
        import fitz  # type: ignore
    except Exception:
        return None, [], None, False

    start = time.monotonic()
    try:
        doc = fitz.open(pdf_path.as_posix())
    except Exception:
        return None, [], None, False

    try:
        total_pages = getattr(doc, "page_count", None)
        if not isinstance(total_pages, int) or total_pages <= 0:
            total_pages = None

        # 扫描范围：最多前 N 页（<=0 表示全量）
        page_limit = total_pages or 0
        if max_pages > 0 and page_limit > 0:
            page_limit = min(page_limit, max_pages)

        has_images_any = False
        render_pages: list[int] = []
        timed_out = False

        for page_index in range(page_limit):
            if max_seconds > 0 and (time.monotonic() - start) >= max_seconds:
                timed_out = True
                break

            page = doc.load_page(page_index)

            # 1) 嵌入位图：扫描/截图/插图等（通常流程图也属于此类）
            try:
                if page.get_images(full=True):
                    has_images_any = True
                    continue
            except Exception:
                # 失败时保守处理：不影响后续其它页扫描
                pass

            # 2) 流程图/矢量图：没有嵌入图片，但存在大量绘制指令（矩形/线条/箭头等）
            try:
                drawings = page.get_drawings() or []
            except Exception:
                drawings = []
            if not drawings:
                continue

            drawings_count = len(drawings)
            if drawings_count >= 20:
                render_pages.append(page_index)
                continue

            # 轻量兜底：绘制不多但文本也很少，仍可能是图示页
            try:
                text = page.get_text("text") or ""
            except Exception:
                text = ""
            if drawings_count >= 5 and len(text.strip()) < 200:
                render_pages.append(page_index)

        return has_images_any, render_pages, total_pages, timed_out
    finally:
        try:
            doc.close()
        except Exception:
            pass


def _extract_pdf_markdown(
    pdf_bytes: bytes,
    filename: str,
    *,
    user_requested_images: bool,
    force_extract_images: bool = False,
    doc_id: str | None = None,
    persisted_pdf_path: Path | None = None,
) -> str:
    """PDF bytes -> Markdown（必要时包含图片/流程图多模态文本）。"""
    images_parser = _get_pdf_images_parser()

    # 说明：
    # - 如果调用方已经把 PDF 落盘（persisted_pdf_path），这里就复用该路径；
    # - 否则走临时目录（仅用于当前解析）。
    temp_dir_ctx = None
    if persisted_pdf_path is None:
        temp_dir_ctx = TemporaryDirectory(prefix="lc_pdf_", ignore_cleanup_errors=True)
        tmp_dir = temp_dir_ctx.__enter__()
        pdf_path = Path(tmp_dir) / _safe_pdf_filename(filename)
        pdf_path.write_bytes(pdf_bytes)
    else:
        pdf_path = persisted_pdf_path

    chunks_path: Path | None = None
    chunks_fp = None
    if doc_id is not None:
        chunks_path, chunks_fp = _open_chunks_writer(
            doc_id,
            filename=filename,
            pdf_path=pdf_path,
        )

    try:
        # 用文件路径打开：更稳定，也避免某些库在 stream 模式下 doc.name=None 的坑
        blob = Blob.from_path(
            str(pdf_path),
            mime_type="application/pdf",
            metadata={"filename": _safe_pdf_filename(filename)},
        )

        # 注意：这里把“落盘分片”和“注入上下文”分开控制：
        # - 分片落盘：尽量全量（默认不限制），避免信息丢失
        # - 注入上下文：严格预算（默认 20 页 / 30000 字符），避免一次请求把模型撑爆
        text_parts: list[str] = []
        context_text_len = 0
        context_text_pages = 0
        extracted_text_len = 0
        extracted_text_pages = 0
        total_text_pages: int | None = None

        for doc in PyMuPDF4LLMParser(mode="page").lazy_parse(blob):
            if total_text_pages is None and isinstance(doc.metadata, dict):
                total_text_pages = doc.metadata.get("total_pages")

            # ==========================
            # chunks.jsonl 是什么？（小白版）
            # ==========================
            # - 我们会把从 PDF 抽取出来的内容写到 storage/pdf_extracted/<DOC_ID>/chunks.jsonl
            # - 文件是 JSON Lines 格式：每一行是一个 JSON（称为一个 chunk）
            # - 每个 chunk 都有一个 kind 字段，用于标记“这段内容从哪里来”：
            #   - kind="text"：PDF 文字层抽取的文本
            #   - kind="images"：PDF 嵌入图片（用豆包多模态识别后得到的文本）
            #   - kind="page_render"：流程图/矢量页（把页面渲染成 PNG，再用豆包识别得到的文本）
            #
            # 后续 pdf_analyze_doc 会逐行读取 chunks.jsonl：
            # - 遇到 images/page_render：直接写入 notes.md（因为豆包已把图变成文本，不再重复调用 deepseek）
            # - 遇到 text：批量喂给 deepseek 生成“增量笔记”（可通过 PDF_ANALYZE_BATCH_SIZE 调整批量大小）
            #
            # 你可以把 chunks.jsonl 理解为：把一份大 PDF 拆成很多“小块内容”，便于断点续跑与按需分析。
            #
            # 按页分片写入（用于后续检索/RAG）
            if chunks_fp is not None:
                parts = _split_text(doc.page_content, _PDF_CHUNK_MAX_CHARS)
                for part_index, part_text in enumerate(parts):
                    chunk = {
                        "doc_id": doc_id,
                        # text：来自 PDF 文本层抽取（不是图片识别）
                        "kind": "text",
                        "page": doc.metadata.get("page")
                        if isinstance(doc.metadata, dict)
                        else None,
                        "part_index": part_index,
                        "part_total": len(parts),
                        "content": part_text,
                        "metadata": doc.metadata,
                    }
                    chunks_fp.write(json.dumps(chunk, ensure_ascii=False) + "\n")

            # 注入上下文的预算：只把“前面的一部分”拼到 text_md 里
            if _PDF_CONTEXT_MAX_PAGES <= 0 or context_text_pages < _PDF_CONTEXT_MAX_PAGES:
                if _PDF_CONTEXT_MAX_CHARS <= 0 or context_text_len < _PDF_CONTEXT_MAX_CHARS:
                    text_parts.append(doc.page_content)
                    context_text_pages += 1
                    context_text_len += len(doc.page_content)

            # 落盘抽取的预算：默认不限制（<=0 表示不限制）
            extracted_text_pages += 1
            extracted_text_len += len(doc.page_content)
            if _PDF_EXTRACT_MAX_PAGES > 0 and extracted_text_pages >= _PDF_EXTRACT_MAX_PAGES:
                break
            if _PDF_EXTRACT_MAX_CHARS > 0 and extracted_text_len >= _PDF_EXTRACT_MAX_CHARS:
                break

        text_md = "\n\n".join(text_parts).strip()
        if total_text_pages and extracted_text_pages < total_text_pages:
            text_md += (
                f"\n\n[提示：文本分片仅落盘解析前 {extracted_text_pages}/{total_text_pages} 页"
                f"（PDF_EXTRACT_MAX_PAGES={_PDF_EXTRACT_MAX_PAGES}，PDF_EXTRACT_MAX_CHARS={_PDF_EXTRACT_MAX_CHARS}）]"
            )
        if total_text_pages and context_text_pages < total_text_pages:
            text_md += (
                f"\n\n[提示：注入到对话上下文的正文为节选：{context_text_pages}/{total_text_pages} 页"
                f"（PDF_CONTEXT_MAX_PAGES={_PDF_CONTEXT_MAX_PAGES}，PDF_CONTEXT_MAX_CHARS={_PDF_CONTEXT_MAX_CHARS}）]"
            )

        mode = (_PDF_IMAGES_MODE or "always").strip().lower()
        # force_extract_images：用于“测试用例工作流”等必须尽量不漏的场景，允许绕过 PDF_IMAGES_MODE=never。
        if mode == "never" and not (_PDF_FORCE_EXTRACT_IMAGES or force_extract_images):
            final_text, _ = _truncate_text(text_md, _PDF_CONTEXT_MAX_CHARS)
            return f"{final_text}\n\n[提示：已关闭PDF图片/流程图解析（PDF_IMAGES_MODE=never）]"

        has_images, render_pages, _total_pages, scan_timed_out = _scan_pdf_visual_content(
            pdf_path,
            max_pages=_PDF_IMAGE_MAX_PAGES,
            max_seconds=_PDF_IMAGE_MAX_SECONDS,
        )
        detected_visuals = (has_images is True) or bool(render_pages)
        scan_failed = has_images is None

        if _PDF_FORCE_EXTRACT_IMAGES:
            enable_images = True
        elif mode == "auto":
            enable_images = user_requested_images or detected_visuals
        else:
            # always / 配置写错：尽量解析（扫描失败时也尝试一次，避免“用户以为能解析，实际没解析”）
            enable_images = user_requested_images or detected_visuals or scan_failed

        if not enable_images:
            final_text, _ = _truncate_text(text_md, _PDF_CONTEXT_MAX_CHARS)
            return (
                f"{final_text}\n\n"
                "[提示：未检测到PDF中的图片/流程图/图表；已仅提取文本。"
                "如需强制尝试解析，请在问题中说明“解析图片”，或设置环境变量 PDF_FORCE_EXTRACT_IMAGES=1]"
            )

        if images_parser is not None:
            try:
                budgeted_parser = _BudgetedImageBlobParser(
                    images_parser,
                    max_images=_PDF_IMAGE_MAX_IMAGES,
                )
                start = time.monotonic()

                visual_sections: list[str] = []
                notes: list[str] = []
                timed_out = False

                # 1) 嵌入图片：用 PyMuPDF4LLMLoader 抽取图片，并由 budgeted_parser（豆包）转为文本
                if has_images is not False:
                    loader = PyMuPDF4LLMLoader(
                        str(pdf_path),
                        mode="page",
                        extract_images=True,
                        images_parser=budgeted_parser,
                    )

                    page_contents: list[str] = []
                    processed_pages = 0
                    total_pages: int | None = None

                    for doc in loader.lazy_load():
                        if total_pages is None and isinstance(doc.metadata, dict):
                            total_pages = doc.metadata.get("total_pages")
                        page_contents.append(doc.page_content)
                        processed_pages += 1

                        if chunks_fp is not None:
                            parts = _split_text(doc.page_content, _PDF_CHUNK_MAX_CHARS)
                            for part_index, part_text in enumerate(parts):
                                chunk = {
                                    "doc_id": doc_id,
                                    # images：来自 PDF 嵌入图片（豆包多模态识别后的文本）
                                    "kind": "images",
                                    "page": doc.metadata.get("page")
                                    if isinstance(doc.metadata, dict)
                                    else None,
                                    "part_index": part_index,
                                    "part_total": len(parts),
                                    "content": part_text,
                                    "metadata": doc.metadata,
                                }
                                chunks_fp.write(json.dumps(chunk, ensure_ascii=False) + "\n")

                        # 约定：<= 0 表示“不限制”
                        if _PDF_IMAGE_MAX_PAGES > 0 and processed_pages >= _PDF_IMAGE_MAX_PAGES:
                            break
                        if _PDF_IMAGE_MAX_SECONDS > 0 and time.monotonic() - start >= _PDF_IMAGE_MAX_SECONDS:
                            timed_out = True
                            break

                    images_md = "\n\n".join(page_contents).strip()
                    if "![" in images_md:
                        visual_sections.append(f"[嵌入图片解析（节选）]\n{images_md}")

                    if total_pages and processed_pages < total_pages:
                        notes.append(
                            f"仅处理前 {processed_pages}/{total_pages} 页（PDF_IMAGE_MAX_PAGES={_PDF_IMAGE_MAX_PAGES}）"
                        )

                # 2) 流程图/矢量图：对特定页做“页面渲染”，再交给 budgeted_parser（豆包）识别
                if render_pages and not timed_out:
                    try:
                        import fitz  # type: ignore
                    except Exception:
                        fitz = None

                    if fitz is not None:
                        rendered_blocks: list[str] = []
                        rendered_pages_count = 0
                        try:
                            doc = fitz.open(pdf_path.as_posix())
                        except Exception:
                            doc = None

                        try:
                            if doc is not None:
                                for page_index in render_pages:
                                    if _PDF_IMAGE_MAX_SECONDS > 0 and time.monotonic() - start >= _PDF_IMAGE_MAX_SECONDS:
                                        timed_out = True
                                        break

                                    page = doc.load_page(page_index)
                                    png_bytes: bytes | None = None
                                    for zoom in (2.0, 1.5, 1.0):
                                        pix = page.get_pixmap(
                                            matrix=fitz.Matrix(zoom, zoom),
                                            alpha=False,
                                        )
                                        candidate = pix.tobytes("png")
                                        if _PDF_IMAGE_MAX_BYTES <= 0 or len(candidate) <= _PDF_IMAGE_MAX_BYTES:
                                            png_bytes = candidate
                                            break

                                    if png_bytes is None:
                                        rendered_blocks.append(
                                            f"[第 {page_index + 1} 页：页面渲染失败/过大，已跳过]"
                                        )
                                        continue

                                    blob = Blob.from_data(
                                        png_bytes,
                                        mime_type="image/png",
                                        path=f"{pdf_path.as_posix()}#page={page_index + 1}",
                                        metadata={
                                            "filename": _safe_pdf_filename(filename),
                                            "page": page_index,
                                        },
                                    )
                                    for img_doc in budgeted_parser.lazy_parse(blob):
                                        text = (img_doc.page_content or "").strip()
                                        if text:
                                            rendered_blocks.append(
                                                f"[第 {page_index + 1} 页：页面渲染解析]\n{text}"
                                            )

                                        if text and chunks_fp is not None and doc_id is not None:
                                            parts = _split_text(text, _PDF_CHUNK_MAX_CHARS)
                                            for part_index, part_text in enumerate(parts):
                                                if not part_text.strip():
                                                    continue
                                                chunk = {
                                                    "doc_id": doc_id,
                                                    # page_render：来自“页面渲染 -> 豆包识别”（用于流程图/矢量图页）
                                                    "kind": "page_render",
                                                    "page": page_index,
                                                    "part_index": part_index,
                                                    "part_total": len(parts),
                                                    "content": part_text,
                                                    "metadata": {"source": pdf_path.as_posix()},
                                                }
                                                chunks_fp.write(json.dumps(chunk, ensure_ascii=False) + "\n")

                                    rendered_pages_count += 1
                                    if budgeted_parser.skipped_images:
                                        break

                        finally:
                            if doc is not None:
                                try:
                                    doc.close()
                                except Exception:
                                    pass

                        rendered_md = "\n\n".join(rendered_blocks).strip()
                        if rendered_md:
                            visual_sections.append(f"[流程图/页面渲染解析（节选）]\n{rendered_md}")
                        if rendered_pages_count:
                            notes.append(f"页面渲染解析 {rendered_pages_count} 页")

                if budgeted_parser.skipped_images:
                    notes.append(
                        f"图片超过上限，已跳过 {budgeted_parser.skipped_images} 张（PDF_IMAGE_MAX_IMAGES={_PDF_IMAGE_MAX_IMAGES}）"
                    )
                if scan_timed_out or timed_out:
                    notes.append(
                        f"达到耗时上限 {int(_PDF_IMAGE_MAX_SECONDS)}s（PDF_IMAGE_MAX_SECONDS={_PDF_IMAGE_MAX_SECONDS}）"
                    )

                combined = text_md
                if visual_sections:
                    combined = (
                        f"{text_md}\n\n---\n\n[图片/流程图多模态解析（节选）]\n"
                        + "\n\n".join(visual_sections)
                    )
                    if notes:
                        combined += "\n\n[" + "；".join(notes) + "]"

                final_text, _ = _truncate_text(combined, _PDF_CONTEXT_MAX_CHARS)
                return final_text
            except Exception as exc:
                final_text, _ = _truncate_text(text_md, _PDF_CONTEXT_MAX_CHARS)
                return f"{final_text}\n\n[图片解析失败，已降级为仅文本提取：{exc}]"
            finally:
                gc.collect()

        final_text, _ = _truncate_text(text_md, _PDF_CONTEXT_MAX_CHARS)
        return final_text
    finally:
        if chunks_fp is not None:
            chunks_fp.close()
        if temp_dir_ctx is not None:
            temp_dir_ctx.__exit__(None, None, None)


def _safe_meta_name(meta: Any) -> str | None:
    if not isinstance(meta, dict):
        return None
    for k in ("filename", "name"):
        v = meta.get(k)
        if isinstance(v, str) and v.strip():
            return v.strip()
    return None


def _describe_image_block(block: dict[str, Any]) -> str:
    block_type = str(block.get("type") or "").lower()
    if block_type == "image_url":
        image_url = block.get("image_url")
        if isinstance(image_url, dict):
            name = _safe_meta_name(image_url.get("metadata"))
            if name:
                return name
    name = _safe_meta_name(block.get("metadata"))
    return name or "uploaded image"


def _data_url_mime_type(url: str) -> str | None:
    if not isinstance(url, str):
        return None
    raw = url.strip()
    if not raw.startswith("data:"):
        return None
    head = raw.split(",", 1)[0]
    # data:<mime_type>;base64
    if not head.startswith("data:"):
        return None
    mt = head[5:].split(";", 1)[0].strip().lower()
    return mt or None


def _try_decode_image_block(block: dict[str, Any]) -> tuple[bytes | None, str, str]:
    """从多模态 block 中提取 (bytes, mime_type, display_name)。失败时 bytes=None。"""
    name = _describe_image_block(block)
    block_type = str(block.get("type") or "").lower().strip()
    mime_type = str(block.get("mime_type") or "").lower().strip() or "image/png"

    raw_b64: str | None = None
    if block_type == "image_url":
        image_url = block.get("image_url")
        if isinstance(image_url, dict):
            url = image_url.get("url")
            if isinstance(url, str) and url.strip():
                url_mt = _data_url_mime_type(url)
                if url_mt:
                    mime_type = url_mt
                # 仅支持 data URL（避免隐式网络请求）
                if url.strip().startswith("data:"):
                    raw_b64 = url
    elif block_type in {"image", "file"}:
        data = block.get("data")
        if isinstance(data, str) and data.strip():
            raw_b64 = data

    if not raw_b64:
        return None, mime_type or "image/png", name

    try:
        return _decode_base64(raw_b64), mime_type or "image/png", name
    except Exception:
        return None, mime_type or "image/png", name


def _replace_pdf_file_blocks_with_text(
    content: Any,
    *,
    force_image_vision: bool = False,
) -> tuple[str | None, bool]:
    """把 HumanMessage.content 中的多模态块替换成可读文本（PDF 抽取 + 图片/附件占位）。"""
    if not isinstance(content, list):
        return None, False

    user_text_parts: list[str] = []
    for block in content:
        if isinstance(block, dict) and block.get("type") == "text":
            text = block.get("text", "")
            if isinstance(text, str) and text.strip():
                user_text_parts.append(text.strip())
    user_text = "\n".join(user_text_parts)

    user_requested_images = _user_requested_images(user_text) or force_image_vision

    text_chunks: list[str] = []
    replaced_any = False

    # 仅在“用户明确要求解析图片”或“强制开启”时才做图片多模态解析，避免普通对话额外开销。
    image_parser: BaseBlobParser | None = None
    if user_requested_images:
        base_parser = _get_pdf_images_parser()
        if base_parser is not None:
            image_parser = _BudgetedImageBlobParser(base_parser, max_images=_PDF_IMAGE_MAX_IMAGES)

    for block in content:
        if isinstance(block, str):
            if block.strip():
                text_chunks.append(block)
            continue

        if not isinstance(block, dict):
            continue

        block_type = block.get("type")
        if block_type == "text":
            text = block.get("text", "")
            if isinstance(text, str) and text.strip():
                text_chunks.append(text)
            continue

        if block_type in {"image_url", "image"}:
            replaced_any = True
            raw, mime_type, name = _try_decode_image_block(block)
            if image_parser is not None and raw is not None:
                blob = Blob.from_data(
                    raw,
                    mime_type=mime_type,
                    path=f"chat://image/{name}",
                    metadata={"filename": name},
                )
                parsed_blocks: list[str] = []
                for img_doc in image_parser.lazy_parse(blob):
                    txt = (img_doc.page_content or "").strip()
                    if txt:
                        parsed_blocks.append(txt)
                parsed = "\n\n".join(parsed_blocks).strip()
                text_chunks.append(f"[图片 {name} 解析结果]\n{parsed}" if parsed else f"[图片已上传：{name}]")
            else:
                text_chunks.append(f"[图片已上传：{name}]")
            continue

        if block_type != "file":
            continue

        mime_type = str(block.get("mime_type") or "").strip().lower()
        filename = _safe_meta_name(block.get("metadata")) or "upload.bin"

        # 兼容 legacy：file + image/*
        if mime_type.startswith("image/"):
            replaced_any = True
            raw, decoded_mime, _name = _try_decode_image_block(
                {
                    "type": "file",
                    "mime_type": mime_type,
                    "data": block.get("data"),
                    "metadata": block.get("metadata"),
                }
            )
            name = filename or _name
            if image_parser is not None and raw is not None:
                blob = Blob.from_data(
                    raw,
                    mime_type=decoded_mime or mime_type,
                    path=f"chat://image/{name}",
                    metadata={"filename": name},
                )
                parsed_blocks: list[str] = []
                for img_doc in image_parser.lazy_parse(blob):
                    txt = (img_doc.page_content or "").strip()
                    if txt:
                        parsed_blocks.append(txt)
                parsed = "\n\n".join(parsed_blocks).strip()
                text_chunks.append(f"[图片 {name} 解析结果]\n{parsed}" if parsed else f"[图片已上传：{name}]")
            else:
                text_chunks.append(f"[图片已上传：{name}]")
            continue

        if (block.get("source_type") or "").lower() != "base64":
            replaced_any = True
            text_chunks.append(f"[附件 {filename}（{mime_type or 'unknown'}）：未提供base64内容，已忽略]")
            continue

        if mime_type != "application/pdf":
            replaced_any = True
            text_chunks.append(f"[附件 {filename}（{mime_type or 'unknown'}）：暂不支持解析，已忽略]")
            continue

        data = block.get("data") or ""
        if not isinstance(data, str) or not data.strip():
            replaced_any = True
            text_chunks.append(f"[附件 {filename}：未提供可用的base64数据]")
            continue

        replaced_any = True
        try:
            pdf_bytes = _decode_base64(data)
            doc_id, persisted_pdf_path = _persist_pdf_upload(pdf_bytes, filename)

            if doc_id is not None and _PDF_REUSE_EXTRACTED and not _PDF_OVERWRITE_EXTRACTED and _extracted_ready(doc_id):
                # 已经抽取过：直接复用落盘分片，避免重复解析导致“从头再来”
                md = _build_context_excerpt_from_chunks(doc_id)
                if not md:
                    md = "[已存在落盘分片，但未能生成节选，请使用 pdf_analyze_doc(DOC_ID, question)]"
            else:
                md = _extract_pdf_markdown(
                    pdf_bytes,
                    filename,
                    user_requested_images=user_requested_images,
                    force_extract_images=force_image_vision,
                    doc_id=doc_id,
                    persisted_pdf_path=persisted_pdf_path,
                )

            # 给用户一个“可追溯 id”：后续要做分片检索/RAG 时，可以用 DOC_ID 定位。
            #
            # 注意：这里不要输出 "doc_id=..."（带等号），否则 system_prompt 可能把它当作“触发条件”，
            # 导致模型在同一轮里反复调用工具，看起来像“一直重复从头解析”。
            if doc_id is not None:
                saved_tip = (
                    f"\n\n[PDF已落盘；DOC_ID: {doc_id}；"
                    f"原始文件目录={Path(_PDF_UPLOAD_DIR).joinpath(doc_id).as_posix()}]"
                )
            else:
                saved_tip = ""

            text_chunks.append(f"[附件 {filename} 解析结果]\n{md}{saved_tip}")
        except Exception as exc:
            text_chunks.append(f"[附件 {filename} 解析失败：{exc}]")

    if not replaced_any:
        return None, False

    return "\n\n".join(text_chunks).strip(), True


def build_pdf_message_updates(
    messages: list[Any],
    *,
    force_image_vision: bool = False,
) -> list[Any]:
    """扫描 messages，把含 PDF/图片等多模态内容的 HumanMessage 更新为“纯文本版”。

    默认行为：仅把图片作为占位符文本（避免普通对话额外开销）。
    当 force_image_vision=True：会尝试把图片内容解析为文本（并且强制开启 PDF 的图片/流程图解析）。
    """
    updated_messages: list[Any] = []

    for msg in messages:
        content = getattr(msg, "content", None)
        new_content, replaced = _replace_pdf_file_blocks_with_text(
            content,
            force_image_vision=force_image_vision,
        )
        if replaced and new_content is not None:
            if getattr(msg, "id", None) is None:
                msg.id = str(uuid4())
            updated_messages.append(msg.model_copy(update={"content": new_content}))

    return updated_messages


def _read_pdf_chunks_jsonl(doc_id: str) -> list[dict[str, Any]]:
    """读取落盘的 chunks.jsonl 并返回 chunks 列表（按写入顺序）。"""
    doc_id = _normalize_doc_id(doc_id)
    chunks_path = Path(_PDF_EXTRACT_DIR) / doc_id / "chunks.jsonl"
    if not chunks_path.exists():
        raise FileNotFoundError(f"未找到分片文件：{chunks_path.as_posix()}")

    chunks: list[dict[str, Any]] = []
    seen: set[str] = set()
    with open(chunks_path, "r", encoding="utf-8") as fp:
        for line in fp:
            line = line.strip()
            if not line:
                continue
            if _PDF_DEDUP_CHUNKS:
                sig = hashlib.sha1(line.encode("utf-8")).hexdigest()
                if sig in seen:
                    continue
                seen.add(sig)
            try:
                obj = json.loads(line)
            except Exception:
                continue
            if isinstance(obj, dict):
                chunks.append(obj)
    return chunks


_DOC_ID_HEX_RE = re.compile(r"[0-9a-fA-F]{64}")


def _normalize_doc_id(value: Any) -> str:
    """把多种“doc_id 传参形式”统一成 64 位 hex（小写）。

    兼容以下输入：
    - 纯哈希：ed8c...（64位）
    - 带前缀：DOC_ID: ed8c...
    - 老格式：doc_id=ed8c...

    这样可以避免：
    - 模型/前端把 DOC_ID 连同前缀一起传进来，导致找不到落盘目录；
    - 用户复制粘贴时带上多余字符，导致“看起来像从头开始”。
    """
    if not isinstance(value, str) or not value.strip():
        raise ValueError("DOC_ID 不能为空，请传入 64 位十六进制哈希。")

    m = _DOC_ID_HEX_RE.search(value.strip())
    if not m:
        raise ValueError(
            "无效 DOC_ID：未在入参中找到 64 位十六进制哈希。"
            "示例：DOC_ID: ed8c506c...（共 64 位）"
        )
    return m.group(0).lower()


@tool
def pdf_analyze_doc(doc_id: str, question: str) -> str:
    """对指定 doc_id 的 PDF 分片做“逐片完整阅读 + 汇总回答”。

    设计目标：
    - 解决“单次上下文塞不下整份 PDF”导致的信息遗漏
    - 用“多次调用模型”逐片处理（每片可控），最后输出尽量完整的解读报告

    关键点：
    - 断点续跑：中途被打断不会从头来（状态落盘在 analysis_state.json）
    - 追加拼接：每次只生成“本分片增量笔记”，直接追加到 notes.md（避免把整份笔记反复塞进模型导致 400）
    - 产物落盘：storage/pdf_extracted/<DOC_ID>/notes.md 与 answer.md
    """
    doc_id = _normalize_doc_id(doc_id)
    model = get_default_model()

    max_steps = _env_int("PDF_ANALYZE_MAX_STEPS", 5000)
    max_seconds = _env_float("PDF_ANALYZE_MAX_SECONDS", 90.0)
    flush_every_steps = _env_int("PDF_ANALYZE_FLUSH_EVERY", 5)
    preview_chars = _env_int("PDF_ANALYZE_NOTES_PREVIEW_CHARS", 1500)
    notes_tail_chars = _env_int("PDF_ANALYZE_NOTES_TAIL_CHARS", 2000)
    # 防御性上限：避免极端情况下写爆磁盘；需要“尽量不漏”可以调大，或设为 0 代表不限。
    notes_max_chars = _env_int("PDF_ANALYZE_NOTES_MAX_CHARS", 300000)

    # ==========================
    # 纯文本分片批处理（减少 deepseek 调用次数）
    # ==========================
    #
    # 小白版解释：
    # - chunks.jsonl 里“纯文本”分片（kind="text"）可能很多。
    # - 原逻辑是“每个分片都调用一次 deepseek 生成增量笔记”，分片多就会调用非常频繁。
    # - 这里做一个“批量合并”：一次把 N 个纯文本分片打包给 deepseek，
    #   让它按分片顺序分别输出增量笔记，从而把调用次数约缩减为 1/N。
    #
    # 如何回退到旧行为？
    # - 设置环境变量：PDF_ANALYZE_BATCH_SIZE=1  （每片一调）
    #
    # 调参建议：
    # - PDF_ANALYZE_BATCH_SIZE：每批多少个分片（默认 4，越大调用越少，但单次提示更长）
    # - PDF_ANALYZE_BATCH_MAX_CHARS：每批“输入正文”最大字符数（默认 16000；<=0 表示不限制）
    batch_size = max(1, _env_int("PDF_ANALYZE_BATCH_SIZE", 4))
    batch_max_chars = _env_int("PDF_ANALYZE_BATCH_MAX_CHARS", 16000)

    final_sections_max_chars = _env_int("PDF_ANALYZE_FINAL_MAX_CHARS", 20000)
    final_input_max_chars = _env_int("PDF_ANALYZE_FINAL_INPUT_MAX_CHARS", 90000)
    chat_return_max_chars = _env_int("PDF_CHAT_RETURN_MAX_CHARS", 6000)

    chunks_path = Path(_PDF_EXTRACT_DIR) / doc_id / "chunks.jsonl"
    if not chunks_path.exists():
        raise FileNotFoundError(f"未找到分片文件：{chunks_path.as_posix()}")

    out_dir = Path(_PDF_EXTRACT_DIR) / doc_id
    out_dir.mkdir(parents=True, exist_ok=True)
    notes_path = out_dir / "notes.md"
    answer_path = out_dir / "answer.md"
    state_path = out_dir / "analysis_state.json"

    def wants_reset(text: str) -> bool:
        """判断用户是否明确想“重新从头分析”。"""
        t = (text or "").strip()
        return any(k in t for k in ("从头", "重来", "重新", "清空", "reset"))

    def _read_tail(path: Path, max_chars: int) -> str:
        """只读文件尾部，避免大文件把内存/上下文撑爆。"""
        if max_chars <= 0 or not path.exists():
            return ""
        try:
            with open(path, "rb") as fp:
                fp.seek(0, os.SEEK_END)
                size = fp.tell()
                fp.seek(max(0, size - 64 * 1024))
                tail_bytes = fp.read()
            tail_text = tail_bytes.decode("utf-8", errors="ignore")
            if len(tail_text) <= max_chars:
                return tail_text.strip()
            return tail_text[-max_chars:].lstrip().strip()
        except Exception:
            return ""

    analysis_goal = (question or "").strip()
    line_offset = 0
    steps = 0
    done = False
    start_ts = time.monotonic()

    if state_path.exists():
        try:
            state = json.loads(state_path.read_text(encoding="utf-8"))
        except Exception:
            state = None
        if isinstance(state, dict):
            line_offset = int(state.get("line_offset", 0) or 0)
            steps = int(state.get("steps", 0) or 0)
            done = bool(state.get("done", False))

            prev_goal = state.get("analysis_goal", "")
            if isinstance(prev_goal, str) and prev_goal.strip() and not done and not wants_reset(question):
                analysis_goal = prev_goal

            # 兼容旧版：旧版把 notes 塞进 state.json，升级后先迁移到 notes.md，避免丢失。
            legacy_notes = state.get("notes")
            if isinstance(legacy_notes, str) and legacy_notes.strip() and not notes_path.exists():
                notes_path.write_text(legacy_notes, encoding="utf-8")

    def flush_state(*, done_flag: bool) -> None:
        payload = {
            "doc_id": doc_id,
            "analysis_goal": analysis_goal,
            "line_offset": line_offset,
            "steps": steps,
            "done": done_flag,
            "updated_at": datetime.now().isoformat(timespec="seconds"),
            "notes_path": notes_path.as_posix(),
            "answer_path": answer_path.as_posix(),
        }
        state_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    # 用户明确要求“重新/从头”：清空 notes + 进度
    if wants_reset(question):
        line_offset = 0
        steps = 0
        done = False
        if notes_path.exists():
            notes_path.write_text("", encoding="utf-8")
        flush_state(done_flag=False)

    system = SystemMessage(
        content=(
            "你是一个严谨的文档分析助手，你将按顺序阅读 PDF 的分片内容。\n"
            "要求：\n"
            "1) 你只输出【本分片新增的关键信息】（增量），不要复述已有内容。\n"
            "2) 增量必须结构化（Markdown 小标题/列表/表格），信息密度高。\n"
            "3) 对规则/优先级/数量限制/字段/埋点/看板等细节要逐条记录。\n"
            "4) 不要照抄原文，要提炼为可用的事实/规则/结论。\n"
        )
    )

    # ==========================
    # 优化点：减少 deepseek 调用次数
    # ==========================
    #
    # 背景（小白版）：
    # - 我们在“PDF 抽取阶段”（见 _extract_pdf_markdown）会把两类视觉信息转成文字：
    #   1) 嵌入图片（kind="images"）：用豆包多模态把图片内容识别为文字
    #   2) 流程图/矢量图（kind="page_render"）：把页面渲染为 PNG，再用豆包多模态识别为文字
    # - 抽取出的文字会被写入 chunks.jsonl（每行一个 chunk），供 pdf_analyze_doc 逐片阅读。
    #
    # 为什么能跳过 deepseek？
    # - 对这些视觉 chunk 来说，“豆包已经把图变成了可检索的文本”，再让 deepseek 逐片做一次“增量总结”
    #   性价比很低：它并不会显著提高最终答案质量，但会增加大量模型调用。
    # - 因此我们做一个条件分支：遇到视觉 chunk 时，直接把它作为“笔记内容”落盘到 notes.md；
    #   deepseek 只在最后“基于累计笔记”生成结构化报告（仍然保证效果一致）。
    #
    # 注意：
    # - 这里不改变“最终汇总/回答”的逻辑，只是避免对同一份视觉内容重复调用 deepseek。
    # - 如果你想恢复原行为，可以把下面 _visual_kinds 设为空集合即可（但不推荐）。
    _visual_kinds = {"images", "page_render"}

    # 防御：notes.md 太大时就停止写入（避免把磁盘写爆/后续读取变慢）。
    # 这个开关只影响“追加笔记”，不影响断点状态（analysis_state.json）的保存。
    notes_write_disabled = False

    def append_notes(
        *,
        chunk_index: Any,
        kind: str,
        page: Any,
        part_index: Any,
        part_total: Any,
        body: str,
    ) -> None:
        nonlocal notes_write_disabled

        # 如果已经触发“停止写入”，后续直接返回（避免重复 I/O）
        if notes_write_disabled:
            return

        body = (body or "").strip()
        if not body:
            return

        # 统一 notes.md 的写入格式：每个 chunk/批次 都带一个小标题 + 元信息（便于人读、也便于后续模型汇总）。
        header = (
            f"\n\n## 分片 {chunk_index}\n"
            f"- kind: {kind}\n"
            f"- page: {page}\n"
            f"- part: {part_index}/{part_total}\n\n"
        )
        to_write = header + body + "\n"

        # 文件大小保护：超过上限就停止继续写入（并写一次提示）。
        # notes_max_chars 是“字节/字符的近似阈值”（用文件大小 stat().st_size 做快速判断）。
        if notes_max_chars > 0 and notes_path.exists():
            try:
                if notes_path.stat().st_size > notes_max_chars:
                    notes_write_disabled = True
                    with open(notes_path, "a", encoding="utf-8") as out_fp:
                        out_fp.write(
                            "\n\n[警告] notes.md 已超过上限，后续增量将不再写入。"
                            "如需继续写入，请调大 PDF_ANALYZE_NOTES_MAX_CHARS 或设置为 0。\n"
                        )
                    return
            except Exception:
                pass

        with open(notes_path, "a", encoding="utf-8") as out_fp:
            out_fp.write(to_write)

    if not done:
        try:
            total_lines = sum(1 for _ in open(chunks_path, "r", encoding="utf-8"))
        except Exception:
            total_lines = None

        seen: set[str] = set()

        # --------------------------
        # 纯文本分片的“批量缓冲区”
        # --------------------------
        # 我们把多个 kind="text" 的分片先放进这个缓冲区，
        # 凑够 batch_size 或达到 batch_max_chars 后，再一次性调用 deepseek。
        text_batch: list[dict[str, Any]] = []
        text_batch_chars = 0

        def flush_text_batch(*, reason: str) -> str | None:
            """把缓冲区里的纯文本分片一次性喂给 deepseek，并把输出写入 notes.md。

            返回：
            - None：成功
            - str：失败时的“已断点保存”提示文本（直接 return 给用户）
            """
            nonlocal text_batch, text_batch_chars, steps, line_offset

            if not text_batch:
                return None

            # 1) 为了去重（避免模型重复输出已有笔记），只把 notes.md 尾部一小段塞进提示词
            notes_tail = _read_tail(notes_path, notes_tail_chars)

            # 2) 构造“多分片输入”，每个分片都带上元信息（page/part），便于模型按分片输出
            batch_blocks: list[str] = []
            for item in text_batch:
                batch_blocks.append(
                    (
                        "----------\n"
                        f"分片 {item.get('chunk_index')}（kind={item.get('kind')} "
                        f"page={item.get('page')} part={item.get('part_index')}/{item.get('part_total')}）\n"
                        f"{item.get('content')}"
                    )
                )
            merged_chunks = "\n\n".join(batch_blocks).strip()

            # 3) 批量提示词：要求模型“按分片顺序分别输出增量”，这样效果尽量接近旧版逐片调用
            prompt = HumanMessage(
                content=(
                    f"分析目标：\n{analysis_goal}\n\n"
                    "你将一次性阅读下面多个分片内容。请【严格按分片顺序】分别输出每个分片的【新增关键信息】（增量）。\n"
                    "输出格式（必须遵守）：\n"
                    "### 分片 <编号>\n"
                    "- ...\n\n"
                    "如果某分片没有新增信息，输出：\n"
                    "### 分片 <编号>\n"
                    "（无新增）\n\n"
                    f"待分析分片（触发原因：{reason}；本批共 {len(text_batch)} 个）：\n"
                    f"{merged_chunks}\n\n"
                    "已记录的累计笔记（末尾节选，仅用于去重，不代表全部）：\n"
                    f"{notes_tail}\n\n"
                    "再次强调：\n"
                    "1) 不要复述笔记节选；\n"
                    "2) 不要跨分片合并输出；\n"
                    "3) 信息密度要高，尽量不漏规则/字段/边界条件。\n"
                )
            )

            try:
                result = model.invoke([system, prompt])
            except Exception as exc:
                # deepseek 调用失败：保存断点并把“可继续指令”返回给用户
                flush_state(done_flag=False)
                preview = _read_tail(notes_path, preview_chars)
                progress = str(line_offset)
                if total_lines:
                    pct = (line_offset / total_lines) * 100
                    progress = f"{line_offset}/{total_lines}（{pct:.1f}%）"
                return (
                    f"[已断点保存] DOC_ID: {doc_id}\n"
                    f"- 已处理到分片行号：{progress}\n"
                    f"- 本轮累计 steps：{steps}\n"
                    f"- 当前进度已落盘：{state_path.as_posix()}\n\n"
                    f"[本次调用遇到错误]\n{exc}\n\n"
                    f"[累计笔记预览]\n{preview}\n\n"
                    f"如需继续，请发送：继续解析 DOC_ID: {doc_id}"
                )

            delta = result.content if isinstance(result.content, str) else str(result.content)

            # 4) 写入 notes.md：用“分片范围”做标题，便于你回看定位
            start_chunk = text_batch[0].get("chunk_index")
            end_chunk = text_batch[-1].get("chunk_index")
            pages = [
                p
                for p in (item.get("page") for item in text_batch)
                if isinstance(p, int)
            ]
            page_hint = f"{min(pages)}..{max(pages)}" if pages else "unknown"

            append_notes(
                chunk_index=f"{start_chunk}-{end_chunk}（批量）",
                kind="text_batch",
                page=page_hint,
                part_index="*",
                part_total="*",
                body=delta,
            )

            # 5) 更新断点状态：
            # - steps：已成功处理的分片数（不是调用次数）
            # - line_offset：已处理到 chunks.jsonl 的哪一行（下次从这里继续）
            steps += len(text_batch)
            last_line_idx = text_batch[-1].get("line_idx")
            try:
                line_offset = int(last_line_idx) + 1
            except Exception:
                # 极端兜底：如果 line_idx 缺失，就不推进（宁可重复，不要跳过）
                pass

            # 按旧逻辑定期落盘进度（可通过 PDF_ANALYZE_FLUSH_EVERY 调整）
            if flush_every_steps > 0 and steps % flush_every_steps == 0:
                flush_state(done_flag=False)

            # 6) 清空缓冲区
            text_batch = []
            text_batch_chars = 0
            return None

        with open(chunks_path, "r", encoding="utf-8") as fp:
            for idx, line in enumerate(fp):
                if idx < line_offset:
                    continue
                # steps 统计“已成功处理的分片数”；text_batch 里的是“已读入但还未处理”的分片
                if max_steps > 0 and (steps + len(text_batch)) >= max_steps:
                    break
                if max_seconds > 0 and (time.monotonic() - start_ts) >= max_seconds:
                    break

                line = line.strip()
                if not line:
                    line_offset = idx + 1
                    continue

                if _PDF_DEDUP_CHUNKS:
                    sig = hashlib.sha1(line.encode("utf-8")).hexdigest()
                    if sig in seen:
                        line_offset = idx + 1
                        continue
                    seen.add(sig)

                try:
                    chunk = json.loads(line)
                except Exception:
                    line_offset = idx + 1
                    continue
                if not isinstance(chunk, dict):
                    line_offset = idx + 1
                    continue

                content = chunk.get("content", "")
                if not isinstance(content, str) or not content.strip():
                    line_offset = idx + 1
                    continue

                kind = chunk.get("kind", "text")
                page = chunk.get("page")
                part_index = chunk.get("part_index")
                part_total = chunk.get("part_total")

                # 统一 kind：避免大小写/空值导致分支判断不稳定
                kind_key = str(kind or "text").strip().lower()

                # 条件边（核心优化）：
                # - 如果是视觉分片（images/page_render），说明内容已经是“豆包识别后的文本”
                # - 直接写入 notes.md 作为“事实笔记”，不要再调用 deepseek 做增量总结
                # - 仍然会推进 line_offset（断点续跑）并按需 flush_state
                if kind_key in _visual_kinds:
                    # 为了保持 notes.md 顺序一致：遇到视觉分片前，先把前面的纯文本缓冲区 flush 掉
                    err = flush_text_batch(reason="before_visual")
                    if err:
                        return err

                    append_notes(
                        chunk_index=idx + 1,
                        kind=kind_key,
                        page=page,
                        part_index=part_index,
                        part_total=part_total,
                        body=content,
                    )
                    steps += 1
                    line_offset = idx + 1
                    if flush_every_steps > 0 and steps % flush_every_steps == 0:
                        flush_state(done_flag=False)
                    continue

                # --------------------------
                # 纯文本分片：进入批量缓冲区
                # --------------------------
                # 说明：
                # - 这里不立刻调用 deepseek，而是先缓存起来；
                # - 缓冲区达到 batch_size 或 batch_max_chars 后，再统一调用一次 deepseek。
                if kind_key != "text":
                    # 兜底：未来如果出现新的 kind（但不属于 _visual_kinds），也按“文本”处理以尽量不漏信息
                    kind_key = "text"

                # 如果“再加一个分片”会超过字符上限，先把当前批次 flush（保持每批大小可控）
                if (
                    text_batch
                    and batch_max_chars > 0
                    and (text_batch_chars + len(content)) > batch_max_chars
                ):
                    err = flush_text_batch(reason="max_chars")
                    if err:
                        return err

                text_batch.append(
                    {
                        "line_idx": idx,
                        "chunk_index": idx + 1,
                        "kind": kind_key,
                        "page": page,
                        "part_index": part_index,
                        "part_total": part_total,
                        "content": content,
                    }
                )
                text_batch_chars += len(content)

                # batch_size==1：相当于旧逻辑（每片一调）；batch_size>1：真正批量
                if batch_size <= 1 or len(text_batch) >= batch_size:
                    err = flush_text_batch(reason="batch_size")
                    if err:
                        return err

            else:
                # for-else：没有 break，说明文件读到结尾
                done = True

        # 文件读完 / 中途 break 后，都尝试 flush 一次残留缓冲区（把已读入的纯文本分片处理完）
        err = flush_text_batch(reason="loop_end")
        if err:
            return err

        if not done:
            try:
                total_lines = sum(1 for _ in open(chunks_path, "r", encoding="utf-8"))
            except Exception:
                total_lines = None
            if total_lines is not None and line_offset >= total_lines:
                done = True

        flush_state(done_flag=done)

        if not done:
            preview = _read_tail(notes_path, preview_chars)
            progress = str(line_offset)
            if total_lines:
                pct = (line_offset / total_lines) * 100
                progress = f"{line_offset}/{total_lines}（{pct:.1f}%）"
            return (
                f"[已断点保存] DOC_ID: {doc_id}\n"
                f"- 已处理到分片行号：{progress}\n"
                f"- 本轮累计 steps：{steps}\n"
                f"- 当前进度已落盘：{state_path.as_posix()}\n\n"
                f"[累计笔记预览]\n{preview}\n\n"
                f"如需继续，请发送：继续解析 DOC_ID: {doc_id}\n"
                f"（提示：可调 PDF_ANALYZE_MAX_SECONDS / PDF_ANALYZE_MAX_STEPS 让单次跑更久）"
            )

    # 已完成：生成最终报告（尽量长，但避免一次输出把前端卡死）
    flush_state(done_flag=True)
    notes = notes_path.read_text(encoding="utf-8") if notes_path.exists() else ""

    final_system = SystemMessage(
        content=(
            "你是一个严谨的PRD/需求文档解读专家。\n"
            "你会收到一份对 PDF 全文逐片阅读后形成的【累计笔记】或其节选。\n"
            "要求：\n"
            "1) 输出要尽可能完整，不要只写一段很短的概括；宁可稍微冗长也不要遗漏。\n"
            "2) 必须结构化：使用 Markdown 标题/表格/列表。\n"
            "3) 对“规则/优先级/数量限制/字段校验/埋点/看板维度”等细节要逐条列出。\n"
            "4) 只基于笔记；如果笔记没有信息，要明确写“笔记未包含该信息”。\n"
        )
    )

    def clamp(text: str) -> str:
        if final_sections_max_chars <= 0:
            return text
        return text[:final_sections_max_chars].rstrip()

    def select_notes_excerpt(all_notes: str, keywords: set[str]) -> str:
        if not all_notes.strip():
            return ""

        paras = [p.strip() for p in all_notes.split("\n\n") if p.strip()]
        head = paras[:40]

        picked: list[str] = []
        for p in paras:
            if any(k in p for k in keywords):
                picked.append(p)

        text = "\n\n".join(head + picked).strip()
        if final_input_max_chars <= 0 or len(text) <= final_input_max_chars:
            return text
        return text[:final_input_max_chars].rstrip()

    section_specs: list[dict[str, Any]] = [
        {
            "task": "请输出【一页纸总览】：需求背景、目标、里程碑、核心方案一句话、范围边界、关键风险。",
            "keywords": {"背景", "目标", "里程碑", "范围", "边界", "风险", "结论"},
        },
        {
            "task": "请输出【竞品分析与结论】：逐个竞品的标签策略、优劣势、对拼拼的启示。",
            "keywords": {"竞品", "叮咚", "盒马", "美团", "拼多多", "对标", "启示", "优劣"},
        },
        {
            "task": "请输出【标签方案（全量细节）】：标签分组、组内/组间优先级、数量限制、标签类型、去掉/新增的标签。",
            "keywords": {"标签", "分组", "优先级", "组内", "组间", "数量", "限制", "规则生产", "手动", "新增", "去掉"},
        },
        {
            "task": "请输出【C端展示规则（必须详细表格）】：商卡页/商详页各位置：位置、标签组、最大数量、优先级、示例；并写变化点。",
            "keywords": {"C端", "商卡", "商详", "腰带", "标题", "位置", "展示", "温层", "促销"},
        },
        {
            "task": "请输出【后台/CMS功能清单】：新增/调整点、审批逻辑、标签管理、优先级(P0)等。",
            "keywords": {"CMS", "后台", "审批", "标签管理", "商卡商详", "P0", "体验优化"},
        },
        {
            "task": "请输出【数据埋点与看板】：埋点列表（页面/位/eid/参数/事件类型）、看板新增维度、数据源、是否实验。",
            "keywords": {"埋点", "曝光", "点击", "eid", "参数", "看板", "维度", "数据源", "实验"},
        },
        {
            "task": "请输出【约束/边界/待确认】：时间约束、范围边界、依赖项、待确认问题。",
            "keywords": {"约束", "依赖", "边界", "待确认", "时间", "风险"},
        },
    ]

    parts: list[str] = []
    for idx, spec in enumerate(section_specs, start=1):
        section_notes = select_notes_excerpt(notes, set(spec.get("keywords") or set()))
        user = HumanMessage(
            content=(
                f"用户问题：\n{question}\n\n"
                f"本章任务（第{idx}章）：\n{spec.get('task')}\n\n"
                f"累计笔记节选（来自 DOC_ID: {doc_id}）：\n{section_notes}"
            )
        )
        resp = model.invoke([final_system, user])
        part_text = resp.content if isinstance(resp.content, str) else str(resp.content)
        parts.append(clamp(part_text))

    answer_text = "\n\n---\n\n".join([p for p in parts if p.strip()]).strip()
    if not answer_text:
        answer_text = "未能生成报告（模型未返回内容），请重试。"

    answer_path.write_text(answer_text, encoding="utf-8")

    chat_text = answer_text
    if chat_return_max_chars > 0 and len(chat_text) > chat_return_max_chars:
        chat_text = (
            chat_text[:chat_return_max_chars].rstrip()
            + "\n\n[提示] 聊天窗口已截断输出，完整报告已落盘，请查看："
            f"{answer_path.as_posix()}"
        )

    return f"{chat_text}\n\n[已落盘：笔记={notes_path.as_posix()}；回答={answer_path.as_posix()}]"


@tool
def pdf_read_report(
    doc_id: str,
    kind: str = "answer",
    offset: int = -1,
    max_chars: int = 6000,
) -> str:
    """分段读取 PDF 的落盘产物（answer.md / notes.md），用于“把完整内容发到聊天里”。

    为什么需要这个工具？
    - answer.md / notes.md 可能很长；一次性把全文塞进聊天返回，容易导致前端卡住、或被网关截断。
    - 这个工具允许你按 offset 分段读取，每次读一段，直到读完。
    - 默认返回“纯内容”，不带状态头/尾，确保前端界面看到的就是报告正文。

    参数说明：
    - doc_id：
      - 支持 `DOC_ID: <hash>` / 纯 hash / 老格式 `doc_id=<hash>`（会自动归一化）
      - 也支持 `auto` / `latest` 或空字符串：自动选择最近一次落盘的 DOC_ID（最省事，但可能选错文档）
    - kind：`answer`（默认）读取最终报告；`notes` 读取累计笔记。
    - offset：从第 offset 个字符开始读取。
      - `-1`（默认）：自动从“上次输出位置”继续（无需用户自己传 offset）。
      - `0`：从头开始输出（重置）。
    - max_chars：本次最多返回多少字符；<=0 表示“尽量全读”（不推荐，容易卡）。
    """
    def auto_pick_latest_doc_id() -> str:
        """自动选择最近一次落盘的 DOC_ID（按文件修改时间）。"""
        base = Path(_PDF_EXTRACT_DIR)
        if not base.exists():
            raise FileNotFoundError(f"未找到目录：{base.as_posix()}")

        best_doc_id: str | None = None
        best_ts: float = -1.0

        for child in base.iterdir():
            if not child.is_dir():
                continue
            name = child.name
            if not _DOC_ID_HEX_RE.fullmatch(name):
                continue

            candidates = [
                child / "answer.md",
                child / "notes.md",
                child / "analysis_state.json",
                child / "report_read_state.json",
            ]
            ts = -1.0
            for p in candidates:
                if p.exists():
                    try:
                        ts = max(ts, p.stat().st_mtime)
                    except Exception:
                        continue
            if ts > best_ts:
                best_ts = ts
                best_doc_id = name

        if best_doc_id is None:
            raise FileNotFoundError(
                f"未发现任何可用的 DOC_ID（{base.as_posix()} 下没有落盘产物）。"
            )
        return best_doc_id

    raw_doc_id = (doc_id or "").strip()
    if not raw_doc_id or raw_doc_id.lower() in {"auto", "latest"}:
        normalized = auto_pick_latest_doc_id()
    else:
        normalized = _normalize_doc_id(raw_doc_id)

    safe_kind = (kind or "").strip().lower()
    if safe_kind not in {"answer", "notes"}:
        return (
            f"kind 参数无效：{kind!r}，仅支持 'answer' 或 'notes'。\n"
            f"示例：pdf_read_report('DOC_ID: {normalized}', kind='answer', offset=0, max_chars=6000)"
        )

    out_dir = Path(_PDF_EXTRACT_DIR) / normalized
    path = out_dir / ("answer.md" if safe_kind == "answer" else "notes.md")
    progress_path = out_dir / "report_read_state.json"
    if not path.exists():
        return (
            f"未找到落盘文件：{path.as_posix()}\n"
            f"请先运行 pdf_analyze_doc('DOC_ID: {normalized}', question) 生成产物。"
        )

    try:
        text = path.read_text(encoding="utf-8")
    except Exception as exc:
        return f"读取失败：{path.as_posix()}：{exc}"

    total = len(text)

    def load_resume_state() -> tuple[int, bool]:
        if not progress_path.exists():
            return 0, False
        try:
            obj = json.loads(progress_path.read_text(encoding="utf-8"))
        except Exception:
            return 0, False
        if not isinstance(obj, dict):
            return 0, False
        by_kind = obj.get(safe_kind)
        if not isinstance(by_kind, dict):
            return 0, False
        v = by_kind.get("offset", 0)
        done = bool(by_kind.get("done", False))
        try:
            return int(v or 0), done
        except Exception:
            return 0, done

    def save_resume_offset(new_offset: int, *, done: bool) -> None:
        payload: dict[str, Any] = {}
        if progress_path.exists():
            try:
                payload = json.loads(progress_path.read_text(encoding="utf-8"))
            except Exception:
                payload = {}
        if not isinstance(payload, dict):
            payload = {}
        payload[safe_kind] = {
            "offset": int(new_offset),
            "done": bool(done),
            "total": int(total),
            "updated_at": datetime.now().isoformat(timespec="seconds"),
        }
        progress_path.write_text(
            json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8"
        )

    try:
        requested_offset = int(offset)
    except Exception:
        requested_offset = -1

    if requested_offset < 0:
        resume_offset, resume_done = load_resume_state()
        # 已经读完且用户没有显式指定 offset：返回空字符串，避免“继续”导致又从头开始。
        if resume_done and resume_offset >= total:
            return ""
        start = resume_offset
    else:
        start = requested_offset

    start = max(0, min(start, total))

    if max_chars is None:
        max_chars = 6000
    try:
        limit = int(max_chars)
    except Exception:
        limit = 6000

    if limit <= 0:
        chunk = text[start:]
        next_offset = total
    else:
        chunk = text[start : start + limit]
        next_offset = start + len(chunk)

    done = next_offset >= total
    save_resume_offset(next_offset, done=done)

    # 只返回“正文内容”，不输出额外状态信息（用户要在前端界面直接看到正文）。
    return chunk
