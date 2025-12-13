import asyncio
import os
from datetime import datetime
from typing import List, Dict, Optional

from langchain_mcp_adapters.client import MultiServerMCPClient
from langchain.agents import create_agent

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    openpyxl = None

try:
    import json
    from pathlib import Path
except ImportError:
    json = None
    Path = None

def get_weather(city:str) ->str:
    """get the weather for a city"""
    return f"{city}.今天是晴天，温度为25度."

def get_zhipu_search_mcp_tools():
    client = MultiServerMCPClient(
        {
            "search": {
                "transport": "sse",  # HTTP-based remote server
                # Ensure you start your weather server on port 8000
                "url": "https://open.bigmodel.cn/api/mcp/web_search/sse?Authorization=dfccd65ce1c9466690f465304cb6ae09.KzH9GpjmAVgN4Jhr",
            }
        }
    )
    tools = asyncio.run(client.get_tools())
    return tools
def get_tavily_search_mcp_tools():
    client = MultiServerMCPClient(
        {
            "search": {
                "transport": "streamable_http",  # HTTP-based remote server
                # Ensure you start your weather server on port 8000
                "url": "https://mcp.tavily.com/mcp/?tavilyApiKey=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",
            }
        }
    )
    tools = asyncio.run(client.get_tools())
    return tools
def get_chrome_mcp_tools():
    client = MultiServerMCPClient(
        {
            "chrome_mcp": {
                "transport": "streamable_http",  # HTTP-based remote server
                # Ensure you start your weather server on port 8000
                "url": "http://127.0.0.1:12306/mcp",
            }
        }
    )
    tools = asyncio.run(client.get_tools())
    return tools
def get_mcp_server_chart_tools():
    client = MultiServerMCPClient(
        {
            "mcp_chart_server": {
                "command":"cmd",
                "args": ["/c", "npx", "-y", "@antv/mcp-server-chart"],
                "transport": "stdio",  # HTTP-based remote server
            }
        }
    )
    tools = asyncio.run(client.get_tools())
    return tools
# tools = get_zhipu_search_mcp_tools()
# print(tools)
# os.environ["TAVILY_API_KEY"] = "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"
# from langchain_tavily import TavilySearch
# toolSearch = TavilySearch(max_results=2)


def save_test_cases_to_excel(
    test_cases: List[Dict],
    file_path: str = "test_cases.xlsx",
    sheet_name: str = "测试用例"
) -> str:
    """
    将测试用例保存到 Excel 文件中

    Args:
        test_cases: 测试用例列表，每个测试用例是一个字典
                   示例: [
                       {
                           "case_id": "TC001",
                           "title": "登录功能测试",
                           "precondition": "用户未登录",
                           "steps": "1. 打开登录页\n2. 输入用户名密码\n3. 点击登录",
                           "expected": "登录成功",
                           "priority": "高",
                           "status": "通过"
                       }
                   ]
        file_path: Excel 文件保存路径，默认为 "test_cases.xlsx"
        sheet_name: 工作表名称，默认为 "测试用例"

    Returns:
        str: 保存结果信息
    """
    # 检查是否安装了 openpyxl 库
    if openpyxl is None:
        return "错误: 未安装 openpyxl 库，请运行: pip install openpyxl"

    # 检查测试用例列表是否为空
    if not test_cases:
        return "错误: 测试用例列表为空"

    try:
        # 创建新的工作簿
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = sheet_name

        # 定义表头（根据测试用例字典的键自动生成）
        headers = list(test_cases[0].keys())

        # 设置表头样式
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 写入表头
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # 写入测试用例数据
        for row_idx, test_case in enumerate(test_cases, start=2):
            for col_idx, header in enumerate(headers, start=1):
                # 获取对应字段的值
                value = test_case.get(header, "")
                cell = sheet.cell(row=row_idx, column=col_idx, value=value)

                # 设置单元格样式
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = border

                # 根据状态设置不同颜色（如果有 status 字段）
                if header.lower() in ["status", "状态"]:
                    if value in ["通过", "Pass", "PASS"]:
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    elif value in ["失败", "Fail", "FAIL"]:
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    elif value in ["阻塞", "Blocked", "BLOCKED"]:
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

        # 自动调整列宽
        for col_idx, header in enumerate(headers, start=1):
            # 计算该列的最大内容长度
            max_length = len(str(header))
            for row_idx in range(2, len(test_cases) + 2):
                cell_value = str(sheet.cell(row=row_idx, column=col_idx).value)
                # 考虑换行符，取最长的一行
                max_line_length = max(len(line) for line in cell_value.split('\n')) if cell_value else 0
                max_length = max(max_length, max_line_length)

            # 设置列宽（限制最大宽度为50）
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = adjusted_width

        # 冻结首行（表头）
        sheet.freeze_panes = "A2"

        # 保存文件
        workbook.save(file_path)

        # 返回成功信息
        return f"成功: 已将 {len(test_cases)} 条测试用例保存到 {os.path.abspath(file_path)}"

    except Exception as e:
        return f"错误: 保存失败 - {str(e)}"


def save_and_generate_report(
    test_cases: List[Dict],
    chart_html: str = "",
    file_path: str = "test_report.html",
    report_title: str = "UI自动化测试报告"
) -> str:
    """
    生成包含图表的HTML测试报告并保存到指定路径

    Args:
        test_cases: 测试用例列表
        chart_html: 图表的HTML代码（从mcp-server-chart生成）
        file_path: 报告保存路径
        report_title: 报告标题

    Returns:
        str: 保存结果信息
    """
    if not test_cases:
        return "错误: 测试用例列表为空"

    try:
        # 统计数据
        total = len(test_cases)
        passed = sum(1 for tc in test_cases if tc.get("状态") in ["通过", "Pass", "PASS"] or tc.get("status") in ["通过", "Pass", "PASS"])
        failed = sum(1 for tc in test_cases if tc.get("状态") in ["失败", "Fail", "FAIL"] or tc.get("status") in ["失败", "Fail", "FAIL"])
        blocked = sum(1 for tc in test_cases if tc.get("状态") in ["阻塞", "Blocked", "BLOCKED"] or tc.get("status") in ["阻塞", "Blocked", "BLOCKED"])
        pass_rate = (passed / total * 100) if total > 0 else 0

        # 生成测试用例表格HTML
        table_rows = ""
        for tc in test_cases:
            status = tc.get("状态") or tc.get("status", "")
            status_class = "pass" if status in ["通过", "Pass", "PASS"] else "fail" if status in ["失败", "Fail", "FAIL"] else "blocked"
            table_rows += f"""
            <tr>
                <td>{tc.get("用例ID") or tc.get("case_id", "")}</td>
                <td>{tc.get("用例标题") or tc.get("title", "")}</td>
                <td>{tc.get("测试步骤") or tc.get("steps", "")}</td>
                <td>{tc.get("预期结果") or tc.get("expected", "")}</td>
                <td>{tc.get("实际结果") or tc.get("actual", "")}</td>
                <td class="{status_class}">{status}</td>
            </tr>
            """

        # HTML模板
        html_content = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{report_title}</title>
    <style>
        body {{ font-family: Arial, sans-serif; margin: 20px; background: #f5f5f5; }}
        .container {{ max-width: 1200px; margin: 0 auto; background: white; padding: 30px; border-radius: 8px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }}
        h1 {{ color: #333; border-bottom: 3px solid #4472C4; padding-bottom: 10px; }}
        .summary {{ display: flex; gap: 20px; margin: 20px 0; }}
        .stat-card {{ flex: 1; padding: 20px; border-radius: 8px; text-align: center; }}
        .stat-card h3 {{ margin: 0; font-size: 32px; }}
        .stat-card p {{ margin: 5px 0 0 0; color: #666; }}
        .total {{ background: #E7F3FF; color: #0066CC; }}
        .pass {{ background: #E8F5E9; color: #2E7D32; }}
        .fail {{ background: #FFEBEE; color: #C62828; }}
        .blocked {{ background: #FFF9E6; color: #F57C00; }}
        .chart-section {{ margin: 30px 0; padding: 20px; background: #fafafa; border-radius: 8px; }}
        table {{ width: 100%; border-collapse: collapse; margin: 20px 0; }}
        th, td {{ padding: 12px; text-align: left; border: 1px solid #ddd; }}
        th {{ background: #4472C4; color: white; font-weight: bold; }}
        tr:nth-child(even) {{ background: #f9f9f9; }}
        .timestamp {{ color: #666; font-size: 14px; margin-top: 20px; text-align: right; }}
    </style>
</head>
<body>
    <div class="container">
        <h1>{report_title}</h1>

        <div class="summary">
            <div class="stat-card total">
                <h3>{total}</h3>
                <p>总用例数</p>
            </div>
            <div class="stat-card pass">
                <h3>{passed}</h3>
                <p>通过</p>
            </div>
            <div class="stat-card fail">
                <h3>{failed}</h3>
                <p>失败</p>
            </div>
            <div class="stat-card blocked">
                <h3>{blocked}</h3>
                <p>阻塞</p>
            </div>
        </div>

        <div class="stat-card" style="background: #F0F4FF; margin: 20px 0;">
            <h3>{pass_rate:.1f}%</h3>
            <p>通过率</p>
        </div>

        {f'<div class="chart-section"><h2>测试统计图表</h2>{chart_html}</div>' if chart_html else ''}

        <h2>测试用例详情</h2>
        <table>
            <thead>
                <tr>
                    <th>用例ID</th>
                    <th>用例标题</th>
                    <th>测试步骤</th>
                    <th>预期结果</th>
                    <th>实际结果</th>
                    <th>状态</th>
                </tr>
            </thead>
            <tbody>
                {table_rows}
            </tbody>
        </table>

        <div class="timestamp">生成时间: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}</div>
    </div>
</body>
</html>"""

        # 保存文件
        Path(file_path).parent.mkdir(parents=True, exist_ok=True)
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(html_content)

        return f"成功: 测试报告已保存到 {os.path.abspath(file_path)}"

    except Exception as e:
        return f"错误: 生成报告失败 - {str(e)}"


# 示例用法
if __name__ == "__main__":
    # 测试数据
    sample_test_cases = [
        {
            "用例ID": "TC001",
            "用例标题": "用户登录功能测试",
            "前置条件": "用户已注册且未登录",
            "测试步骤": "1. 打开登录页面\n2. 输入正确的用户名和密码\n3. 点击登录按钮",
            "预期结果": "登录成功，跳转到首页",
            "优先级": "高",
            "状态": "通过",
            "创建时间": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        },
        {
            "用例ID": "TC002",
            "用例标题": "用户登录失败测试",
            "前置条件": "用户已注册且未登录",
            "测试步骤": "1. 打开登录页面\n2. 输入错误的密码\n3. 点击登录按钮",
            "预期结果": "提示密码错误",
            "优先级": "高",
            "状态": "失败",
            "创建时间": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        },
        {
            "用例ID": "TC003",
            "用例标题": "购物车添加商品测试",
            "前置条件": "用户已登录",
            "测试步骤": "1. 浏览商品列表\n2. 点击添加到购物车\n3. 查看购物车",
            "预期结果": "商品成功添加到购物车",
            "优先级": "中",
            "状态": "阻塞",
            "创建时间": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }
    ]

    # 调用函数保存测试用例
    result = save_test_cases_to_excel(sample_test_cases, "测试用例报告.xlsx")
    print(result)